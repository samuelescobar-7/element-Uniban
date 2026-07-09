import re
from io import BytesIO
from pathlib import Path
from tempfile import NamedTemporaryFile
from datetime import datetime
from zoneinfo import ZoneInfo

import math

import openpyxl
import pandas as pd
import streamlit as st
import streamlit.components.v1 as components


# =========================
# CONFIGURACIÓN
# =========================
COL_RESPUESTA_F = 6
COL_RESPUESTA_G = 7
COL_RESPUESTA_K = 11

VALID_RESPUESTAS_F = {
    "SÍ (ESTÁNDAR ERP)",
    "SI (COMPONENTE ADICIONAL)",
    "DES",
    "TER",
    "NO",
}
VALID_RESPUESTAS_G = {"SI", "NO"}
VALID_RESPUESTAS_K = {"COMPLETA", "CASI COMPLETA", "PARCIALMENTE COMPLETA", "INCOMPLETA", "TOTALMENTE INCOMPLETA"}

HOJAS_EXCLUIDAS = {"1.31", "1.32", "1.33"}
COL_RESPUESTA_NF_D = 4
COL_RESPUESTA_NF_E = 5
VALID_RESPUESTAS_NF_D = {"SI", "NO"}
VALID_RESPUESTAS_NF_E = {"SI", "NO"}

# =========================
# "OTRAS" — filas adicionales de % en blanco (hoja "% en blanco")
# =========================
# Cada entrada define: etiqueta a mostrar en la columna "Proceso", el prefijo de la
# hoja de origen, el rango de filas a inspeccionar y el número de columna donde
# están las respuestas. El % en blanco se calcula como:
#   (celdas vacías en ese rango) / (total de filas del rango) * 100
OTRAS_FILAS_BLANCO_CONFIG = [
    {
        "label": "2. Información de la Solución",
        "hoja_prefijo": "2.",
        "fila_inicio": 5,
        "fila_fin": 28,
        "columna": 3,  # Columna C
    },
]

# =========================
# "OTRAS" — filas adicionales de conteo de registros (hoja "% en blanco")
# =========================
# Cada entrada define: etiqueta a mostrar en la columna "Proceso" y la clave con la
# que se encuentra la lista de DataFrames (uno por proveedor) en el diccionario de
# fuentes de datos que se pasa a `construir_tabla_conteo_registros`. El conteo es
# simplemente el número de filas (registros) consolidados por proveedor en esa
# fuente; si un proveedor no tiene registros, se reporta 0.
OTRAS_CONTEO_REGISTROS_CONFIG = [
    {
        "label": "Numero de registros en 3. Experiencia del Fabricante",
        "fuente": "data_experiencia_raw",
    },
    {
        "label": "Numero de registros en 5. Experiencia del Proponente",
        "fuente": "data_experiencia_oferente_raw",
    },
]

# =========================
# "OTRAS" — filas adicionales de % en blanco de la columna C (Incluido SI/NO)
# para "7. Alcance Servicios" y "8. Metodología"
# =========================
# Cada entrada define: etiqueta a mostrar en la columna "Proceso" y la clave con la
# que se encuentra la lista de DataFrames (uno por proveedor, con la columna
# "Respuesta_C") ya calculada previamente para Alcance/Metodología
# (data_alcance_servicios / data_metodologia). El % en blanco se calcula sobre el
# mismo total de filas (100%) usado en esos cálculos, por proveedor.
OTRAS_BLANCOS_COLUMNA_C_CONFIG = [
    {
        "label": "7. Alcance Servicios",
        "fuente": "data_alcance_servicios",
    },
    {
        "label": "8. Metodología",
        "fuente": "data_metodologia",
    },
]

# =========================
# AGRUPACIÓN DE HOJAS FUNCIONALES (1.X)
# =========================
GRUPOS_HOJAS_FUNC = [
    ("COMERCIALIZADOR INTERNACIONAL", ["1.1", "1.4", "1.5", "1.6", "1.34"]),
    ("LOGÍSTICA Y AGENCIA DE CARGA", ["1.7.1", "1.7.2", "1.8", "1.9", "1.10"]),
    ("PRODUCCIÓN FRUTÍCOLA Y FINCAS ALIADAS - GESTIÓN FINCAS", ["1.2", "1.3", "1.28"]),
    ("FÁBRICAS CAJAS", ["1.13"]),
    ("FÁBRICAS PLASTICOS", ["1.14"]),
    ("NEGOCIO SNACKS", ["1.15"]),
    ("PRODUCTOS Y SERVICIOS AL CAMPO", ["1.16", "1.17"]),
    ("TRANSVERSALES A LOS NEGOCIOS - CALIDAD Y TRAZABILIDAD", ["1.11", "1.12"]),
    ("SERVICIOS CORPORATIVOS", ["1.18", "1.19", "1.20", "1.21", "1.22", "1.23", "1.24", "1.25", "1.26", "1.27"]),
    ("FUNDACIÓN", ["1.29"]),
    ("TMA EUROPA", ["1.30"]),
]

MARCADOR_GRUPO = "▶ "


def obtener_prefijo_hoja(nombre_hoja):
    nombre = str(nombre_hoja).strip()
    m = re.match(r"^(\d+\.\d+(?:\.\d+)?)", nombre)
    return m.group(1) if m else None


def obtener_grupo_hoja(nombre_hoja):
    """Devuelve el nombre del grupo/categoría al que pertenece una hoja 1.X, o None si no aplica."""
    prefijo = obtener_prefijo_hoja(nombre_hoja)
    if prefijo is None:
        return None
    for grupo_nombre, prefijos in GRUPOS_HOJAS_FUNC:
        if prefijo in prefijos:
            return grupo_nombre
    return None


def es_fila_grupo(valor_hoja):
    """True si la fila corresponde a un encabezado de grupo (no a una hoja real)."""
    return isinstance(valor_hoja, str) and valor_hoja.startswith(MARCADOR_GRUPO)


def agrupar_df_por_categoria(df, columna_hoja="Hoja"):
    """
    Reordena las filas de un DataFrame indexado por hoja, insertando una fila de
    encabezado (vacía salvo el nombre del grupo) antes de las hojas funcionales 1.X
    que pertenezcan a cada categoría, en el orden de GRUPOS_HOJAS_FUNC.
    Las hojas no reconocidas (p.ej. TOTAL u otras no funcionales) se dejan al final,
    en su orden original, sin encabezado de grupo.
    """
    if df is None or df.empty or columna_hoja not in df.columns:
        return df

    columnas = list(df.columns)
    otras_cols = [c for c in columnas if c != columna_hoja]

    grupos_por_fila = df[columna_hoja].apply(obtener_grupo_hoja)

    filas_finales = []
    hojas_usadas = set()

    for grupo_nombre, _ in GRUPOS_HOJAS_FUNC:
        mascara = grupos_por_fila == grupo_nombre
        sub_df = df[mascara]
        if sub_df.empty:
            continue
        fila_header = {columna_hoja: f"{MARCADOR_GRUPO}{grupo_nombre}"}
        for c in otras_cols:
            fila_header[c] = ""
        filas_finales.append(fila_header)
        for _, row in sub_df.iterrows():
            filas_finales.append({c: row[c] for c in columnas})
            hojas_usadas.add(row[columna_hoja])

    resto = df[~df[columna_hoja].isin(hojas_usadas)]
    for _, row in resto.iterrows():
        filas_finales.append({c: row[c] for c in columnas})

    return pd.DataFrame(filas_finales, columns=columnas)


# =========================
# FUNCIONES
# =========================
def normalizar(valor):
    if valor is None:
        return ""
    return str(valor).strip().upper()


def es_hoja_1x(nombre):
    nombre = nombre.strip()
    prefijo = re.match(r"^(\d+\.\d+)", nombre)
    if prefijo and prefijo.group(1) in HOJAS_EXCLUIDAS:
        return False
    return bool(re.match(r"^1\.", nombre))


def es_hoja_no_funcional(nombre):
    nombre = nombre.strip()
    prefijo = re.match(r"^(\d+\.\d+)", nombre)
    if prefijo and prefijo.group(1) in HOJAS_EXCLUIDAS:
        return True
    return False


def detectar_filas(ws):
    filas = []
    for r in range(2, ws.max_row + 1):
        val = normalizar(ws.cell(r, 1).value)
        if "*** FIN DEL DOCUMENTO ***" in val:
            break
        if val.isdigit():
            filas.append(r)
    return filas


def leer_respuesta(ws, fila, col, validas):
    val = ws.cell(fila, col).value
    if val is None:
        return "VACIO"
    val = normalizar(val)
    if val in validas:
        return val
    return "VACIO"


def analizar_hoja(ws, pesos_f, peso_col_f, peso_col_g):
    data = []
    for r in detectar_filas(ws):
        resp_f = leer_respuesta(ws, r, COL_RESPUESTA_F, VALID_RESPUESTAS_F)
        resp_g = leer_respuesta(ws, r, COL_RESPUESTA_G, VALID_RESPUESTAS_G)
        id_req = ws.cell(r, 1).value
        requerimiento = ws.cell(r, 5).value
        peso_f = pesos_f.get(resp_f, 0.0)
        peso_g = peso_col_g if resp_g == "SI" else 0.0
        data.append({
            "Hoja": ws.title,
            "Fila": r,
            "ID": str(id_req).strip() if id_req is not None else "",
            "Requerimiento": requerimiento,
            "Resp_F": resp_f,
            "Resp_G": resp_g,
            "Peso_F": peso_f,
            "Peso_G": peso_g,
            "Peso_Total": peso_f + peso_g
        })
    df = pd.DataFrame(data)
    if df.empty:
        return None, None
    maximo_posible = peso_col_f + peso_col_g
    if maximo_posible == 0:
        return 0.0, df
    cumplimiento = (df["Peso_Total"].mean() / maximo_posible) * 100
    return round(cumplimiento, 2), df


def analizar_hoja_k(ws, pesos_k, col_requerimiento=5):
    data = []
    for r in detectar_filas(ws):
        resp_k = leer_respuesta(ws, r, COL_RESPUESTA_K, VALID_RESPUESTAS_K)
        id_req = ws.cell(r, 1).value
        requerimiento = ws.cell(r, col_requerimiento).value
        peso_k = pesos_k.get(resp_k, 0.0)
        data.append({
            "Hoja": ws.title,
            "Fila": r,
            "ID": str(id_req).strip() if id_req is not None else "",
            "Requerimiento": requerimiento,
            "Peso_K": peso_k
        })
    df = pd.DataFrame(data)
    if df.empty:
        return None, None
    maximo_posible = pesos_k.get("COMPLETA", 1.0)
    if maximo_posible == 0:
        return 0.0, df
    calidad = (df["Peso_K"].mean() / maximo_posible) * 100
    return round(calidad, 2), df


def analizar_hoja_nf(ws, peso_col_d, peso_col_e):
    data = []
    for r in detectar_filas(ws):
        resp_d = leer_respuesta(ws, r, COL_RESPUESTA_NF_D, VALID_RESPUESTAS_NF_D)
        resp_e = leer_respuesta(ws, r, COL_RESPUESTA_NF_E, VALID_RESPUESTAS_NF_E)
        id_req = ws.cell(r, 1).value
        requerimiento = ws.cell(r, 3).value
        peso_d = peso_col_d if resp_d == "SI" else 0.0
        peso_e = peso_col_e if resp_e == "SI" else 0.0
        data.append({
            "Hoja": ws.title,
            "Fila": r,
            "ID": str(id_req).strip() if id_req is not None else "",
            "Requerimiento": requerimiento,
            "Resp_D": resp_d,
            "Resp_E": resp_e,
            "Peso_F": peso_d,
            "Peso_G": peso_e,
            "Peso_Total": peso_d + peso_e
        })
    df = pd.DataFrame(data)
    if df.empty:
        return None, None
    maximo_posible = peso_col_d + peso_col_e
    if maximo_posible == 0:
        return 0.0, df
    cumplimiento = (df["Peso_Total"].mean() / maximo_posible) * 100
    return round(cumplimiento, 2), df


def analizar_archivo(path, pesos_f, peso_col_f, peso_col_g, pesos_k):
    wb = openpyxl.load_workbook(path, data_only=True)

    hojas_func = [s for s in wb.sheetnames if es_hoja_1x(s)]
    resultados, detalles, resultados_k, detalles_k = {}, {}, {}, {}
    for h in hojas_func:
        ws = wb[h]
        cumplimiento, detalle_df = analizar_hoja(ws, pesos_f, peso_col_f, peso_col_g)
        if cumplimiento is not None:
            resultados[h] = cumplimiento
            detalles[h] = detalle_df
        calidad, detalle_k_df = analizar_hoja_k(ws, pesos_k)
        if calidad is not None:
            resultados_k[h] = calidad
            detalles_k[h] = detalle_k_df

    hojas_nofunc = [s for s in wb.sheetnames if es_hoja_no_funcional(s)]
    resultados_nf, detalles_nf, resultados_k_nf, detalles_k_nf = {}, {}, {}, {}
    for h in hojas_nofunc:
        ws = wb[h]
        cumplimiento, detalle_df = analizar_hoja_nf(ws, peso_col_f, peso_col_g)
        if cumplimiento is not None:
            resultados_nf[h] = cumplimiento
            detalles_nf[h] = detalle_df
        calidad, detalle_k_df = analizar_hoja_k(ws, pesos_k, col_requerimiento=3)
        if calidad is not None:
            resultados_k_nf[h] = calidad
            detalles_k_nf[h] = detalle_k_df

    return (resultados, detalles, resultados_k, detalles_k,
            resultados_nf, detalles_nf, resultados_k_nf, detalles_k_nf)


def analizar_hoja_experiencia_raw(wb, proveedor):
    COLUMNAS = [
        "Proveedor",
        "Negocio",
        "Nombre de la empresa (Cliente)",
        "País donde se realizó la implementación",
        "Nombre del contacto",
        "E-mail del contacto",
        "Página Web",
        "Nombre del Producto Instalado y Funcionando",
    ]

    hoja_nombre = next((s for s in wb.sheetnames if s.strip().startswith("3.")), None)
    if hoja_nombre is None:
        return pd.DataFrame(columns=COLUMNAS)

    ws = wb[hoja_nombre]

    merged_values = {}
    for merge_range in ws.merged_cells.ranges:
        top_left = ws.cell(merge_range.min_row, merge_range.min_col).value
        for row in range(merge_range.min_row, merge_range.max_row + 1):
            for col in range(merge_range.min_col, merge_range.max_col + 1):
                merged_values[(row, col)] = top_left

    def get_cell_value(r, c):
        if (r, c) in merged_values:
            return merged_values[(r, c)]
        return ws.cell(r, c).value

    ENCABEZADOS = {
        "nombre de la empresa", "país donde", "nombre del contacto",
        "e-mail", "página web", "nombre del producto", "#"
    }

    filas_procesadas = set()
    data = []
    negocio_actual = ""

    for r in range(1, ws.max_row + 1):
        val_a = ws.cell(r, 1).value

        es_numero = False
        if val_a is not None:
            try:
                int(str(val_a).strip())
                es_numero = True
            except ValueError:
                pass

        if es_numero:
            if r in filas_procesadas:
                continue
            filas_procesadas.add(r)

            valores = [ws.cell(r, c).value for c in range(2, 8)]
            if all(v is None or str(v).strip() == "" for v in valores):
                continue

            fila = {
                "Proveedor": proveedor,
                "Negocio": negocio_actual,
                "Nombre de la empresa (Cliente)":              str(get_cell_value(r, 2) or "").strip(),
                "País donde se realizó la implementación":     str(get_cell_value(r, 3) or "").strip(),
                "Nombre del contacto":                         str(get_cell_value(r, 4) or "").strip(),
                "E-mail del contacto":                         str(get_cell_value(r, 5) or "").strip(),
                "Página Web":                                  str(get_cell_value(r, 6) or "").strip(),
                "Nombre del Producto Instalado y Funcionando": str(get_cell_value(r, 7) or "").strip(),
            }
            data.append(fila)

        else:
            texto_fila = ""
            for c in range(1, ws.max_column + 1):
                v = get_cell_value(r, c)
                if v is not None and str(v).strip():
                    texto_fila = str(v).strip()
                    break

            if not texto_fila:
                continue

            texto_lower = texto_fila.lower()

            if any(enc in texto_lower for enc in ENCABEZADOS):
                continue

            if len(texto_fila) < 5:
                continue

            negocio_actual = texto_fila

    if not data:
        return pd.DataFrame(columns=COLUMNAS)
    return pd.DataFrame(data)[COLUMNAS]


# =========================
# EXPERIENCIA OFERENTE — resumen pivot por industria
# =========================
def analizar_hoja_experiencia_oferente(wb, proveedor):
    hoja_nombre = next((s for s in wb.sheetnames if s.strip().startswith("5.")), None)
    if hoja_nombre is None:
        return pd.DataFrame([{"Sector/Industria": "", "País": "", "Proveedor": proveedor}])
    ws = wb[hoja_nombre]
    data = []
    for r in range(9, ws.max_row + 1):
        num = ws.cell(r, 2).value
        if num is None:
            continue
        sector = ws.cell(r, 7).value
        pais   = ws.cell(r, 3).value
        if sector is None and pais is None:
            continue
        data.append({
            "Sector/Industria": str(sector).strip() if sector else "",
            "País": str(pais).strip() if pais else "",
            "Proveedor": proveedor
        })
    if not data:
        return pd.DataFrame([{"Sector/Industria": "", "País": "", "Proveedor": proveedor}])
    return pd.DataFrame(data)


# =========================
# EXPERIENCIA OFERENTE — detalle completo (raw)
# =========================
def analizar_hoja_experiencia_oferente_raw(wb, proveedor):
    COLUMNAS = [
        "Nombre del contratante (Cliente)",
        "País donde se realizó la implementación",
        "Nombre del contacto",
        "E-mail del contacto",
        "Página Web",
        "Tipo de Industria",
        "Procesos implementados",
        "Describa las integraciones del sistema core (ERP) y las soluciones avanzadas",
    ]
    COL_INICIO = 2
    COL_FIN    = 9

    hoja_nombre = next((s for s in wb.sheetnames if s.strip().startswith("5.")), None)
    if hoja_nombre is None:
        return pd.DataFrame(columns=["Proveedor"] + COLUMNAS)

    ws = wb[hoja_nombre]
    data = []
    for r in range(9, ws.max_row + 1):
        valores = [ws.cell(r, c).value for c in range(COL_INICIO, COL_FIN + 1)]
        if all(v is None for v in valores):
            continue
        fila = {"Proveedor": proveedor}
        for col_name, val in zip(COLUMNAS, valores):
            fila[col_name] = str(val).strip() if val is not None else ""
        data.append(fila)

    if not data:
        return pd.DataFrame(columns=["Proveedor"] + COLUMNAS)
    return pd.DataFrame(data)[["Proveedor"] + COLUMNAS]


# =========================
# INFORMACIÓN DE LA SOLUCIÓN — Diferenciadores tecnicos
# =========================
def analizar_hoja_info_solucion(wb, proveedor):
    hoja_nombre = next((s for s in wb.sheetnames if s.strip().startswith("2.")), None)
    if hoja_nombre is None:
        return pd.DataFrame(columns=["Proveedor", "Requerimiento", "Respuesta"])
    ws = wb[hoja_nombre]
    FILAS = [21]
    data = []
    for r in FILAS:
        req_val = ws.cell(r, 2).value
        res_val = ws.cell(r, 3).value
        data.append({
            "Proveedor":     proveedor,
            "Requerimiento": str(req_val).strip() if req_val is not None else "",
            "Respuesta":     str(res_val).strip() if res_val is not None else "",
        })
    return pd.DataFrame(data)[["Proveedor", "Requerimiento", "Respuesta"]]


# =========================
# INFORMACIÓN DE LA SOLUCIÓN — Evolución
# =========================
def analizar_hoja_evolucion(wb, proveedor):
    hoja_nombre = next((s for s in wb.sheetnames if s.strip().startswith("2.")), None)
    if hoja_nombre is None:
        return pd.DataFrame(columns=["Proveedor", "Requerimiento", "Respuesta"])
    ws = wb[hoja_nombre]
    FILAS = [28]
    data = []
    for r in FILAS:
        req_val = ws.cell(r, 2).value
        res_val = ws.cell(r, 3).value
        data.append({
            "Proveedor":     proveedor,
            "Requerimiento": str(req_val).strip() if req_val is not None else "",
            "Respuesta":     str(res_val).strip() if res_val is not None else "",
        })
    return pd.DataFrame(data)[["Proveedor", "Requerimiento", "Respuesta"]]


# =========================
# INFORMACIÓN DE LA SOLUCIÓN — red de partners, mecanismos de soporte y mantenimiento
# =========================
def analizar_hoja_ecosistema(wb, proveedor):
    hoja_nombre = next((s for s in wb.sheetnames if s.strip().startswith("2.")), None)
    if hoja_nombre is None:
        return pd.DataFrame(columns=["Proveedor", "Requerimiento", "Respuesta"])
    ws = wb[hoja_nombre]
    FILAS = [23, 27]
    data = []
    for r in FILAS:
        req_val = ws.cell(r, 2).value
        res_val = ws.cell(r, 3).value
        data.append({
            "Proveedor":     proveedor,
            "Requerimiento": str(req_val).strip() if req_val is not None else "",
            "Respuesta":     str(res_val).strip() if res_val is not None else "",
        })
    return pd.DataFrame(data)[["Proveedor", "Requerimiento", "Respuesta"]]


# =========================
# INFORMACIÓN DE LA SOLUCIÓN — Centros de (I+D)
# =========================
def analizar_hoja_centros_id(wb, proveedor):
    hoja_nombre = next((s for s in wb.sheetnames if s.strip().startswith("2.")), None)
    if hoja_nombre is None:
        return pd.DataFrame(columns=["Proveedor", "Requerimiento", "Respuesta"])
    ws = wb[hoja_nombre]
    FILAS = [24]
    data = []
    for r in FILAS:
        req_val = ws.cell(r, 2).value
        res_val = ws.cell(r, 3).value
        data.append({
            "Proveedor":     proveedor,
            "Requerimiento": str(req_val).strip() if req_val is not None else "",
            "Respuesta":     str(res_val).strip() if res_val is not None else "",
        })
    return pd.DataFrame(data)[["Proveedor", "Requerimiento", "Respuesta"]]


# =========================
# INFORMACIÓN DE LA SOLUCIÓN — Comunidades colaborativas
# =========================
def analizar_hoja_comunidades(wb, proveedor):
    hoja_nombre = next((s for s in wb.sheetnames if s.strip().startswith("2.")), None)
    if hoja_nombre is None:
        return pd.DataFrame(columns=["Proveedor", "Requerimiento", "Respuesta"])
    ws = wb[hoja_nombre]
    FILAS = [25]
    data = []
    for r in FILAS:
        req_val = ws.cell(r, 2).value
        res_val = ws.cell(r, 3).value
        data.append({
            "Proveedor":     proveedor,
            "Requerimiento": str(req_val).strip() if req_val is not None else "",
            "Respuesta":     str(res_val).strip() if res_val is not None else "",
        })
    return pd.DataFrame(data)[["Proveedor", "Requerimiento", "Respuesta"]]


# =========================
# ALCANCE DE SERVICIOS
# =========================
def analizar_hoja_alcance_servicios(wb, proveedor):
    hoja_nombre = next((s for s in wb.sheetnames if s.strip().startswith("7.")), None)
    if hoja_nombre is None:
        return pd.DataFrame(columns=["Respuesta_C", "Respuesta_E", "Proveedor"])
    ws = wb[hoja_nombre]
    data = []
    for r in range(6, ws.max_row + 1):
        nombre = ws.cell(r, 2).value
        if nombre is None or str(nombre).strip() == "":
            continue
        val_c = ws.cell(r, 3).value
        val_c_norm = str(val_c).strip().upper() if val_c is not None else "VACIO"
        if val_c_norm not in {"SI", "NO"}:
            val_c_norm = "VACIO"
        val_e = ws.cell(r, 5).value
        val_e_norm = str(val_e).strip().upper() if val_e is not None else "VACIO"
        if val_e_norm not in VALID_RESPUESTAS_K:
            val_e_norm = "VACIO"
        data.append({
            "Respuesta_C": val_c_norm,
            "Respuesta_E": val_e_norm,
            "Proveedor": proveedor
        })
    if not data:
        return pd.DataFrame(columns=["Respuesta_C", "Respuesta_E", "Proveedor"])
    return pd.DataFrame(data)


def analizar_hoja_alcance_servicios_raw(wb, proveedor):
    COLUMNAS = [
        "Servicio",
        "Incluido (SI/NO)",
        "Explicación o Descripción Adicional",
        "Calidad",
    ]
    COL_INICIO = 2
    COL_FIN = 5

    hoja_nombre = next((s for s in wb.sheetnames if s.strip().startswith("7.")), None)
    if hoja_nombre is None:
        return pd.DataFrame(columns=["Proveedor"] + COLUMNAS)

    ws = wb[hoja_nombre]
    data = []
    for r in range(6, ws.max_row + 1):
        valores = [ws.cell(r, c).value for c in range(COL_INICIO, COL_FIN + 1)]
        if all(v is None for v in valores):
            continue
        fila = {"Proveedor": proveedor}
        for col_name, val in zip(COLUMNAS, valores):
            fila[col_name] = str(val).strip() if val is not None else ""
        data.append(fila)

    if not data:
        return pd.DataFrame(columns=["Proveedor"] + COLUMNAS)
    return pd.DataFrame(data)[["Proveedor"] + COLUMNAS]


def analizar_hoja_metodologia(wb, proveedor):
    hoja_nombre = next((s for s in wb.sheetnames if s.strip().startswith("8.")), None)
    if hoja_nombre is None:
        return pd.DataFrame(columns=["Respuesta_C", "Respuesta_E", "Proveedor"])
    ws = wb[hoja_nombre]
    data = []
    for r in range(6, ws.max_row + 1):
        nombre = ws.cell(r, 2).value
        if nombre is None or str(nombre).strip() == "":
            continue
        val_c = ws.cell(r, 3).value
        val_c_norm = str(val_c).strip().upper() if val_c is not None else "VACIO"
        if val_c_norm not in {"SI", "NO"}:
            val_c_norm = "VACIO"
        val_e = ws.cell(r, 5).value
        val_e_norm = str(val_e).strip().upper() if val_e is not None else "VACIO"
        if val_e_norm not in VALID_RESPUESTAS_K:
            val_e_norm = "VACIO"
        data.append({
            "Respuesta_C": val_c_norm,
            "Respuesta_E": val_e_norm,
            "Proveedor": proveedor
        })
    if not data:
        return pd.DataFrame(columns=["Respuesta_C", "Respuesta_E", "Proveedor"])
    return pd.DataFrame(data)


def analizar_hoja_metodologia_raw(wb, proveedor):
    COLUMNAS = [
        "Elemento de la Metodología",
        "Incluido (SI/NO)",
        "Explicación o Descripción Adicional (Especificar página de la propuesta si existe información adicional sobre este ítem en ella)",
        "Calidad",
    ]
    COL_INICIO = 2
    COL_FIN = 5

    hoja_nombre = next((s for s in wb.sheetnames if s.strip().startswith("8.")), None)
    if hoja_nombre is None:
        return pd.DataFrame(columns=["Proveedor"] + COLUMNAS)

    ws = wb[hoja_nombre]
    data = []
    for r in range(6, ws.max_row + 1):
        valores = [ws.cell(r, c).value for c in range(COL_INICIO, COL_FIN + 1)]
        if all(v is None for v in valores):
            continue
        fila = {"Proveedor": proveedor}
        for col_name, val in zip(COLUMNAS, valores):
            fila[col_name] = str(val).strip() if val is not None else ""
        data.append(fila)

    if not data:
        return pd.DataFrame(columns=["Proveedor"] + COLUMNAS)
    return pd.DataFrame(data)[["Proveedor"] + COLUMNAS]


# =========================
# CONSTRUCCIÓN DE TABLAS
# =========================
def construir_tablas_cumplimiento(data):
    df = pd.DataFrame(data)
    orden_hojas = list(dict.fromkeys(df["Hoja"].tolist()))
    df_final = (
        df.pivot_table(
            index="Hoja", columns="Proveedor", values="Cumplimiento_%", aggfunc="first"
        )
        .fillna(0)
        .reindex(orden_hojas)
        .reset_index()
    )
    df_final.columns.name = None
    total_por_proveedor = df.groupby("Proveedor")["Cumplimiento_%"].mean().round(2).to_dict()
    df_total = pd.DataFrame([{"Hoja": "TOTAL", **total_por_proveedor}]).fillna(0)
    return df_final, df_total


def construir_tablas_calidad(data_k):
    df_k = pd.DataFrame(data_k)
    orden_hojas = list(dict.fromkeys(df_k["Hoja"].tolist()))
    df_final_k = (
        df_k.pivot_table(
            index="Hoja", columns="Proveedor", values="Calidad_%", aggfunc="first"
        )
        .fillna(0)
        .reindex(orden_hojas)
        .reset_index()
    )
    df_final_k.columns.name = None
    total_k_por_proveedor = df_k.groupby("Proveedor")["Calidad_%"].mean().round(2).to_dict()
    df_total_k = pd.DataFrame([{"Hoja": "TOTAL", **total_k_por_proveedor}]).fillna(0)
    return df_final_k, df_total_k


def construir_tabla_integrada(
    df_final_cum, df_final_k,
    pesos_hojas,
    peso_total_cumplimiento, peso_total_calidad
):
    proveedores = [c for c in df_final_cum.columns if c != "Hoja"]
    df_cum = df_final_cum.set_index("Hoja")
    df_cal = df_final_k.set_index("Hoja") if df_final_k is not None else None

    hojas = df_cum.index.tolist()
    filas = []
    for h in hojas:
        fila = {"Hoja": h}
        peso_h = pesos_hojas.get(h, 100) / 100
        for prov in proveedores:
            cum = df_cum.loc[h, prov] if prov in df_cum.columns else 0.0
            cal = df_cal.loc[h, prov] if (df_cal is not None and h in df_cal.index and prov in df_cal.columns) else 0.0
            fila[prov] = round(
                ((cum * peso_total_cumplimiento) + (cal * peso_total_calidad)) * peso_h,
                2
            )
        filas.append(fila)

    df_integrado = pd.DataFrame(filas)
    total_fila = {"Hoja": "TOTAL"}
    for prov in proveedores:
        total_fila[prov] = round(df_integrado[prov].sum(), 2)
    df_total_integrado = pd.DataFrame([total_fila])
    return df_integrado, df_total_integrado


def formatear_porcentaje_df(df):
    df_fmt = df.copy()
    for col in df_fmt.columns:
        if col != "Hoja":
            df_fmt[col] = df_fmt[col].apply(
                lambda x: f"{x:.2f}%" if isinstance(x, (int, float)) and not isinstance(x, bool) else x
            )
    return df_fmt


def df_to_excel_bytes(dfs: dict) -> bytes:
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for sheet_name, df in dfs.items():
            df.to_excel(writer, index=False, sheet_name=sheet_name)
    return buf.getvalue()


def boton_descarga(label, dfs: dict, file_name: str, key: str):
    st.download_button(
        label,
        df_to_excel_bytes(dfs),
        file_name=file_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=key
    )


def obtener_fecha_modificacion(archivo_bytes):
    import zipfile
    import xml.etree.ElementTree as ET
    import re as _re

    try:
        with zipfile.ZipFile(BytesIO(archivo_bytes)) as z:
            with z.open("docProps/core.xml") as f:
                tree = ET.parse(f)
                root = tree.getroot()
                ns = {"dcterms": "http://purl.org/dc/terms/"}
                modified_el = root.find("dcterms:modified", ns)
                if modified_el is None or not modified_el.text:
                    return "No disponible"

                raw = modified_el.text.strip()
                formatos = [
                    "%Y-%m-%dT%H:%M:%SZ",
                    "%Y-%m-%dT%H:%M:%S",
                    "%Y-%m-%d %H:%M:%S",
                    "%Y-%m-%dT%H:%M",
                    "%Y-%m-%d",
                ]
                dt = None
                for fmt in formatos:
                    try:
                        dt = datetime.strptime(raw, fmt).replace(tzinfo=ZoneInfo("UTC"))
                        break
                    except ValueError:
                        continue

                if dt is None:
                    try:
                        clean = _re.sub(r"Z$", "+00:00", raw)
                        dt = datetime.fromisoformat(clean)
                        if dt.tzinfo is None:
                            dt = dt.replace(tzinfo=ZoneInfo("UTC"))
                    except ValueError:
                        pass

                if dt is not None:
                    dt_local = dt.astimezone(ZoneInfo("America/Bogota"))
                    return dt_local.strftime("%Y-%m-%d %H:%M:%S")

    except Exception:
        pass
    return "No disponible"


def construir_hoja_info_analisis(
    fecha_generacion,
    peso_col_f,
    peso_col_g,
    pesos_f,
    pesos_k,
    peso_total_cumplimiento,
    peso_total_calidad,
    pesos_hojas_func,
    pesos_hojas_nf,
    metadata_archivos,
    peso_alcance,
    peso_metodologia,
):
    bloques = []

    df_fecha = pd.DataFrame([{"Fecha y hora de generación del reporte": fecha_generacion}])
    bloques.append(("Generación del reporte", df_fecha))

    params_generales = [
        {"Parámetro": "Peso máximo columna F (cubrimiento)", "Valor": peso_col_f},
        {"Parámetro": "Peso máximo columna G (inclusión)",   "Valor": peso_col_g},
    ]
    df_params = pd.DataFrame(params_generales)
    bloques.append(("Parámetros generales", df_params))

    pesos_f_rows = [{"Respuesta": k, "Peso aplicado": v} for k, v in pesos_f.items()]
    df_pesos_f = pd.DataFrame(pesos_f_rows)
    bloques.append(("Pesos cubrimiento (columna F)", df_pesos_f))

    pesos_k_rows = [{"Respuesta": k, "Peso aplicado": v} for k, v in pesos_k.items()]
    df_pesos_k = pd.DataFrame(pesos_k_rows)
    bloques.append(("Pesos calidad (columna K)", df_pesos_k))

    df_pesos_totales = pd.DataFrame([
        {"Parámetro": "Peso total cumplimiento", "Valor": peso_total_cumplimiento},
        {"Parámetro": "Peso total calidad",      "Valor": peso_total_calidad},
    ])
    bloques.append(("Pesos totales puntaje combinado", df_pesos_totales))

    if pesos_hojas_func:
        df_ph_func = pd.DataFrame([{"Hoja": h, "Peso asignado (%)": p} for h, p in pesos_hojas_func.items()])
        bloques.append(("Pesos por hoja funcional", df_ph_func))

    if pesos_hojas_nf:
        df_ph_nf = pd.DataFrame([{"Hoja": h, "Peso asignado (%)": p} for h, p in pesos_hojas_nf.items()])
        bloques.append(("Pesos por hoja no funcional", df_ph_nf))

    df_pesos_adicionales = pd.DataFrame([
        {"Parámetro": "Peso alcance",     "Valor": peso_alcance},
        {"Parámetro": "Peso metodología", "Valor": peso_metodologia},
    ])
    bloques.append(("Pesos alcance y metodología", df_pesos_adicionales))

    if metadata_archivos:
        df_archivos = pd.DataFrame(metadata_archivos)
        bloques.append(("Archivos analizados", df_archivos))

    return bloques


def escribir_hoja_info_analisis(writer, bloques):
    ws = writer.book.create_sheet("Informacion de analisis")
    fila_actual = 1

    for titulo, df in bloques:
        cell = ws.cell(row=fila_actual, column=1, value=titulo)
        cell.font = openpyxl.styles.Font(bold=True, size=11)
        fila_actual += 1

        for col_idx, col_name in enumerate(df.columns, start=1):
            ws.cell(row=fila_actual, column=col_idx, value=col_name).font = \
                openpyxl.styles.Font(bold=True)
        fila_actual += 1

        for _, row in df.iterrows():
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=fila_actual, column=col_idx, value=value)
            fila_actual += 1

        fila_actual += 1

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)


def _safe_to_excel(df, writer, sheet_name):
    if df is not None and not df.empty:
        df.to_excel(writer, index=False, sheet_name=sheet_name)


def _orden_requerimientos(lista_dfs):
    df_all = pd.concat(lista_dfs).sort_values("Fila")
    pares = list(zip(df_all["ID"].tolist(), df_all["Requerimiento"].tolist()))
    return list(dict.fromkeys(pares))


def _pivot_ordenado(df_datos, col_valor, orden):
    df_idx = df_datos.copy()
    df_idx["_idx"] = list(zip(df_idx["ID"], df_idx["Requerimiento"]))
    pivot = (
        df_idx.pivot_table(
            index="_idx",
            columns="Proveedor",
            values=col_valor,
            aggfunc="first"
        )
        .fillna(0)
        .reindex(orden)
        .reset_index()
    )
    pivot.columns.name = None
    pivot.insert(0, "ID", pivot["_idx"].apply(lambda t: t[0]))
    pivot.insert(1, "Requerimiento", pivot["_idx"].apply(lambda t: t[1]))
    pivot = pivot.drop(columns=["_idx"])
    return pivot


def pivotar_requerimiento_proveedor(df, nombres_proveedores=None):
    """
    Convierte un DataFrame en formato largo (Proveedor | Requerimiento | Respuesta)
    en un DataFrame en formato ancho: Requerimiento | Proveedor1 | Proveedor2 | ... | ProveedorN.
    El número de filas resultantes corresponde al número de requerimientos (filas)
    analizados en la hoja de origen, preservando el orden original en que aparecen.
    """
    if df is None or df.empty:
        return df

    orden_req = list(dict.fromkeys(df["Requerimiento"].tolist()))

    pivot = (
        df.pivot_table(
            index="Requerimiento",
            columns="Proveedor",
            values="Respuesta",
            aggfunc="first"
        )
        .reindex(orden_req)
        .reset_index()
    )
    pivot.columns.name = None
    pivot = pivot.fillna("")

    if nombres_proveedores:
        cols = ["Requerimiento"] + [p for p in nombres_proveedores if p in pivot.columns]
        pivot = pivot[cols]

    return pivot


def calcular_tabla_alcance(data_alcance_servicios, nombres_proveedores, pesos_k,
                            peso_total_cumplimiento, peso_total_calidad, peso_alcance):
    if not data_alcance_servicios:
        return None, None

    df_alc_all = pd.concat(data_alcance_servicios, ignore_index=True)
    todos_provs = nombres_proveedores if nombres_proveedores else sorted(df_alc_all["Proveedor"].unique())
    max_k = pesos_k.get("COMPLETA", 1.0)

    fila_fmt = {"Métrica": "Puntaje alcance"}
    fila_raw = {"Métrica": "Puntaje alcance"}

    for prov in todos_provs:
        df_prov = df_alc_all[df_alc_all["Proveedor"] == prov]
        total = len(df_prov)
        if total == 0:
            fila_fmt[prov] = "0.00%"
            fila_raw[prov] = 0.0
            continue

        n_si = (df_prov["Respuesta_C"] == "SI").sum()
        pct_si = (n_si / total) * 100

        if max_k > 0:
            pct_cal = df_prov["Respuesta_E"].map(lambda v: pesos_k.get(v, 0.0)).mean() / max_k * 100
        else:
            pct_cal = 0.0

        puntaje = round(
            (pct_si * peso_total_cumplimiento + pct_cal * peso_total_calidad) * peso_alcance, 2
        )
        fila_fmt[prov] = f"{puntaje:.2f}%"
        fila_raw[prov] = puntaje

    return pd.DataFrame([fila_fmt]), pd.DataFrame([fila_raw])


def calcular_tabla_metodologia(data_metodologia, nombres_proveedores, pesos_k,
                                peso_total_cumplimiento, peso_total_calidad, peso_metodologia):
    if not data_metodologia:
        return None, None

    df_met_all = pd.concat(data_metodologia, ignore_index=True)
    todos_provs = nombres_proveedores if nombres_proveedores else sorted(df_met_all["Proveedor"].unique())
    max_k = pesos_k.get("COMPLETA", 1.0)

    fila_fmt = {"Métrica": "Puntaje metodología"}
    fila_raw = {"Métrica": "Puntaje metodología"}

    for prov in todos_provs:
        df_prov = df_met_all[df_met_all["Proveedor"] == prov]
        total = len(df_prov)
        if total == 0:
            fila_fmt[prov] = "0.00%"
            fila_raw[prov] = 0.0
            continue

        n_si = (df_prov["Respuesta_C"] == "SI").sum()
        pct_si = (n_si / total) * 100

        if max_k > 0:
            pct_cal = df_prov["Respuesta_E"].map(lambda v: pesos_k.get(v, 0.0)).mean() / max_k * 100
        else:
            pct_cal = 0.0

        puntaje = round(
            (pct_si * peso_total_cumplimiento + pct_cal * peso_total_calidad) * peso_metodologia, 2
        )
        fila_fmt[prov] = f"{puntaje:.2f}%"
        fila_raw[prov] = puntaje

    return pd.DataFrame([fila_fmt]), pd.DataFrame([fila_raw])


def analizar_blancos_rango(wb, hoja_prefijo, fila_inicio, fila_fin, columna):
    """
    Calcula cuántas celdas están en blanco en una columna específica, dentro de un
    rango de filas fijo, en la primera hoja cuyo nombre empiece con `hoja_prefijo`.
    Devuelve (blancos, total) o (None, None) si no se encuentra la hoja.
    """
    hoja_nombre = next((s for s in wb.sheetnames if s.strip().startswith(hoja_prefijo)), None)
    if hoja_nombre is None:
        return None, None
    ws = wb[hoja_nombre]
    total = fila_fin - fila_inicio + 1
    blancos = 0
    for r in range(fila_inicio, fila_fin + 1):
        val = ws.cell(r, columna).value
        if val is None or str(val).strip() == "":
            blancos += 1
    return blancos, total


def construir_tabla_blancos_otras(data_otras_blancos, nombres_proveedores):
    """
    Construye la tabla de la sección "Otras" de la hoja "% en blanco", a partir de
    los conteos recolectados por `analizar_blancos_rango` para cada entrada de
    OTRAS_FILAS_BLANCO_CONFIG.

    Devuelve un DataFrame con columnas: Proceso | <proveedor1> | <proveedor2> | ...
    """
    if not data_otras_blancos:
        return None

    filas = []
    for cfg in OTRAS_FILAS_BLANCO_CONFIG:
        label = cfg["label"]
        fila = {"Proceso": label}
        conteos_label = data_otras_blancos.get(label, {})
        for prov in nombres_proveedores:
            conteo = conteos_label.get(prov)
            if conteo is None:
                fila[prov] = 0.0
                continue
            blancos, total = conteo
            if blancos is None or not total:
                fila[prov] = 0.0
            else:
                fila[prov] = round((blancos / total) * 100, 2)
        filas.append(fila)

    columnas_orden = ["Proceso"] + nombres_proveedores
    df_otras = pd.DataFrame(filas)
    for c in columnas_orden:
        if c not in df_otras.columns:
            df_otras[c] = 0.0
    df_otras = df_otras[columnas_orden]
    return df_otras


# =========================
# "OTRAS" — conteo de registros consolidados (no es %, es cantidad)
# =========================
def construir_tabla_conteo_registros(config_list, fuentes_datos, nombres_proveedores):
    """
    Construye filas de conteo de registros (cantidades, no porcentajes) para la
    sección "Otras" de la hoja "% en blanco".

    `fuentes_datos` es un diccionario donde cada clave coincide con el campo
    "fuente" de `config_list` y el valor es la lista de DataFrames (uno o más por
    proveedor, tal como se acumulan durante el procesamiento, p.ej.
    `data_experiencia_raw` o `data_experiencia_oferente_raw`) usada para construir
    las hojas "Exp - Fabricante completa" y "Exp - Oferente completa".

    El conteo por proveedor es simplemente el número de filas (registros)
    consolidados para ese proveedor en la fuente correspondiente; si el proveedor
    no tiene registros, se reporta 0.

    Devuelve un DataFrame con columnas: Proceso | <proveedor1> | <proveedor2> | ...
    """
    if not config_list:
        return None

    filas = []
    for cfg in config_list:
        label = cfg["label"]
        lista_dfs = fuentes_datos.get(cfg["fuente"]) or []

        conteos = {}
        if lista_dfs:
            df_all = pd.concat(lista_dfs, ignore_index=True)
            if not df_all.empty and "Proveedor" in df_all.columns:
                conteos = df_all.groupby("Proveedor").size().to_dict()

        fila = {"Proceso": label}
        for prov in nombres_proveedores:
            fila[prov] = int(conteos.get(prov, 0))
        filas.append(fila)

    columnas_orden = ["Proceso"] + nombres_proveedores
    df_conteo = pd.DataFrame(filas)
    for c in columnas_orden:
        if c not in df_conteo.columns:
            df_conteo[c] = 0
    df_conteo = df_conteo[columnas_orden]
    return df_conteo


# =========================
# "OTRAS" — % en blanco de la columna C (Incluido SI/NO) para
# "7. Alcance Servicios" y "8. Metodología"
# =========================
def construir_tabla_blancos_columna_c(config_list, fuentes_datos, nombres_proveedores):
    """
    Construye filas de % en blanco de la columna C ("Incluido (SI/NO)", almacenada
    como "Respuesta_C") para la sección "Otras" de la hoja "% en blanco", a partir
    de las listas de DataFrames ya calculadas para Alcance de servicios y
    Metodología (las mismas usadas en `calcular_tabla_alcance` y
    `calcular_tabla_metodologia`).

    El % en blanco se calcula, por proveedor, como:
        (filas con Respuesta_C == "VACIO") / (total de filas del proveedor) * 100
    usando el mismo total (100%) que ya se usa en el cálculo del puntaje de
    alcance/metodología.

    Devuelve un DataFrame con columnas: Proceso | <proveedor1> | <proveedor2> | ...
    """
    if not config_list:
        return None

    filas = []
    for cfg in config_list:
        label = cfg["label"]
        lista_dfs = fuentes_datos.get(cfg["fuente"]) or []
        df_all = pd.concat(lista_dfs, ignore_index=True) if lista_dfs else pd.DataFrame()

        fila = {"Proceso": label}
        for prov in nombres_proveedores:
            if df_all.empty or "Proveedor" not in df_all.columns:
                fila[prov] = 0.0
                continue
            df_prov = df_all[df_all["Proveedor"] == prov]
            total = len(df_prov)
            if total == 0 or "Respuesta_C" not in df_prov.columns:
                fila[prov] = 0.0
                continue
            blancos = (df_prov["Respuesta_C"] == "VACIO").sum()
            fila[prov] = round((blancos / total) * 100, 2)
        filas.append(fila)

    columnas_orden = ["Proceso"] + nombres_proveedores
    df_res = pd.DataFrame(filas)
    for c in columnas_orden:
        if c not in df_res.columns:
            df_res[c] = 0.0
    df_res = df_res[columnas_orden]
    return df_res


# =========================
# % EN BLANCO — Requerimientos funcionales
# =========================
def construir_tabla_blancos_funcional(detalles_globales, nombres_proveedores):
    """
    Construye una tabla con el % de casillas en blanco (VACIO) de la columna F
    y de la columna G, por hoja funcional (proceso) y por proveedor, calculado
    sobre el total de filas reconocidas en el cumplimiento para ese proveedor
    en esa hoja.

    Devuelve un DataFrame con columnas:
    Proceso | Columna F <prov1> | Columna G <prov1> | ... | Columna F <provN> | Columna G <provN>
    """
    if not detalles_globales:
        return None

    filas = []
    for hoja, lista_dfs in detalles_globales.items():
        fila = {"Proceso": hoja}
        for prov in nombres_proveedores:
            df_prov = None
            for df_ in lista_dfs:
                if df_ is not None and not df_.empty and df_["Proveedor"].iloc[0] == prov:
                    df_prov = df_
                    break

            col_f_name = f"Columna F {prov}"
            col_g_name = f"Columna G {prov}"

            if df_prov is None or df_prov.empty or "Resp_F" not in df_prov.columns:
                fila[col_f_name] = 0.0
                fila[col_g_name] = 0.0
                continue

            total = len(df_prov)
            blancos_f = (df_prov["Resp_F"] == "VACIO").sum()
            blancos_g = (df_prov["Resp_G"] == "VACIO").sum()

            fila[col_f_name] = round((blancos_f / total) * 100, 2) if total else 0.0
            fila[col_g_name] = round((blancos_g / total) * 100, 2) if total else 0.0

        filas.append(fila)

    columnas_orden = ["Proceso"]
    for prov in nombres_proveedores:
        columnas_orden.append(f"Columna F {prov}")
        columnas_orden.append(f"Columna G {prov}")

    df_blancos = pd.DataFrame(filas)
    for c in columnas_orden:
        if c not in df_blancos.columns:
            df_blancos[c] = 0.0
    df_blancos = df_blancos[columnas_orden]
    return df_blancos


def construir_tabla_blancos_no_funcional(detalles_globales_nf, nombres_proveedores):
    """
    Igual que construir_tabla_blancos_funcional, pero para las hojas no funcionales,
    donde las respuestas relevantes están en las columnas D y E (Resp_D / Resp_E).

    Devuelve un DataFrame con columnas:
    Proceso | Columna D <prov1> | Columna E <prov1> | ... | Columna D <provN> | Columna E <provN>
    """
    if not detalles_globales_nf:
        return None

    filas = []
    for hoja, lista_dfs in detalles_globales_nf.items():
        fila = {"Proceso": hoja}
        for prov in nombres_proveedores:
            df_prov = None
            for df_ in lista_dfs:
                if df_ is not None and not df_.empty and df_["Proveedor"].iloc[0] == prov:
                    df_prov = df_
                    break

            col_d_name = f"Columna D {prov}"
            col_e_name = f"Columna E {prov}"

            if df_prov is None or df_prov.empty or "Resp_D" not in df_prov.columns:
                fila[col_d_name] = 0.0
                fila[col_e_name] = 0.0
                continue

            total = len(df_prov)
            blancos_d = (df_prov["Resp_D"] == "VACIO").sum()
            blancos_e = (df_prov["Resp_E"] == "VACIO").sum()

            fila[col_d_name] = round((blancos_d / total) * 100, 2) if total else 0.0
            fila[col_e_name] = round((blancos_e / total) * 100, 2) if total else 0.0

        filas.append(fila)

    columnas_orden = ["Proceso"]
    for prov in nombres_proveedores:
        columnas_orden.append(f"Columna D {prov}")
        columnas_orden.append(f"Columna E {prov}")

    df_blancos = pd.DataFrame(filas)
    for c in columnas_orden:
        if c not in df_blancos.columns:
            df_blancos[c] = 0.0
    df_blancos = df_blancos[columnas_orden]
    return df_blancos


def invertir_porcentaje_df(df, filas_sin_invertir=None):
    """
    Devuelve una copia de un DataFrame de tipo "% en blanco" (columna "Proceso" +
    columnas numéricas por proveedor) donde cada valor numérico se invierte como
    100 - valor, para obtener el "% Diligenciado" equivalente.

    Las filas cuyo valor de "Proceso" esté en `filas_sin_invertir` (por ejemplo,
    las filas de conteo de registros, que son cantidades y no porcentajes) se
    dejan sin cambios.
    """
    if df is None or df.empty:
        return df

    filas_sin_invertir = filas_sin_invertir or set()
    df_inv = df.copy()

    for idx, row in df_inv.iterrows():
        if row.get("Proceso") in filas_sin_invertir:
            continue
        for col in df_inv.columns:
            if col == "Proceso":
                continue
            val = row[col]
            if isinstance(val, (int, float)) and not isinstance(val, bool):
                df_inv.at[idx, col] = round(100 - val, 2)

    return df_inv


def _escribir_bloque_blancos(ws, df_bloque, titulo, columnas_orden, fila_inicio, agrupar=False, filas_enteras=None):
    """Escribe un bloque (título + encabezados + filas) de % en blanco a partir de la fila_inicio.
    Las filas cuyo valor de "Proceso" esté en `filas_enteras` se escriben como números
    enteros (conteos), sin formato de porcentaje.
    Devuelve la siguiente fila libre después del bloque."""
    fila_actual = fila_inicio
    n_cols = len(columnas_orden)
    filas_enteras = filas_enteras or set()

    ws.cell(row=fila_actual, column=1, value=titulo).font = openpyxl.styles.Font(bold=True, size=12)
    if n_cols > 1:
        ws.merge_cells(start_row=fila_actual, start_column=1, end_row=fila_actual, end_column=n_cols)
    fila_actual += 1

    if df_bloque is None or df_bloque.empty:
        ws.cell(row=fila_actual, column=1, value="No se encontraron datos.")
        return fila_actual + 2

    for col_idx, col_name in enumerate(columnas_orden, start=1):
        ws.cell(row=fila_actual, column=col_idx, value=col_name).font = openpyxl.styles.Font(bold=True)
    fila_actual += 1

    df_escribir = agrupar_df_por_categoria(df_bloque, columna_hoja="Proceso") if agrupar else df_bloque

    for _, row in df_escribir.iterrows():
        es_fila_entera = row.get("Proceso") in filas_enteras
        for col_idx, col_name in enumerate(columnas_orden, start=1):
            val = row.get(col_name, "")
            if col_name != "Proceso" and isinstance(val, (int, float)) and not isinstance(val, bool):
                val = int(val) if es_fila_entera else f"{val:.2f}%"
            ws.cell(row=fila_actual, column=col_idx, value=val)
        fila_actual += 1

    return fila_actual + 1


def _sanitizar_nombre_hoja(nombre, existentes):
    """
    Convierte el nombre de un proveedor en un nombre de hoja Excel válido:
    quita caracteres no permitidos (: \\ / ? * [ ]), lo recorta a 31 caracteres
    y evita colisiones con nombres de hoja ya usados en `existentes` (se
    modifica in-place agregando el nombre final).
    """
    invalidos = set(':\\/?*[]')
    limpio = "".join(c for c in str(nombre) if c not in invalidos).strip()
    if not limpio:
        limpio = "Proveedor"
    limpio = limpio[:31]

    base = limpio
    contador = 1
    while limpio in existentes:
        sufijo = f"_{contador}"
        limpio = base[: 31 - len(sufijo)] + sufijo
        contador += 1

    existentes.add(limpio)
    return limpio


def construir_bloques_diligenciado_proveedor(df_blancos_func_dilig, df_blancos_nf_dilig,
                                              df_blancos_otras_dilig, proveedor):
    """
    Extrae, para un único proveedor, los tres bloques de "% Diligenciado"
    (funcional, no funcional y "Otras") a partir de las tablas consolidadas
    (que tienen una columna por proveedor), devolviendo DataFrames de una sola
    columna de valores (además de "Proceso"), con nombres de columna genéricos
    ("Columna F", "Columna G", "Columna D", "Columna E", "% Diligenciado")
    listos para escribirse en la hoja propia del proveedor.
    """
    df_func_prov = None
    if df_blancos_func_dilig is not None and not df_blancos_func_dilig.empty:
        f_col = f"Columna F {proveedor}"
        g_col = f"Columna G {proveedor}"
        cols = ["Proceso"]
        rename = {}
        if f_col in df_blancos_func_dilig.columns:
            cols.append(f_col)
            rename[f_col] = "Columna F"
        if g_col in df_blancos_func_dilig.columns:
            cols.append(g_col)
            rename[g_col] = "Columna G"
        if len(cols) > 1:
            df_func_prov = df_blancos_func_dilig[cols].rename(columns=rename)

    df_nf_prov = None
    if df_blancos_nf_dilig is not None and not df_blancos_nf_dilig.empty:
        d_col = f"Columna D {proveedor}"
        e_col = f"Columna E {proveedor}"
        cols = ["Proceso"]
        rename = {}
        if d_col in df_blancos_nf_dilig.columns:
            cols.append(d_col)
            rename[d_col] = "Columna D"
        if e_col in df_blancos_nf_dilig.columns:
            cols.append(e_col)
            rename[e_col] = "Columna E"
        if len(cols) > 1:
            df_nf_prov = df_blancos_nf_dilig[cols].rename(columns=rename)

    df_otras_prov = None
    if (df_blancos_otras_dilig is not None and not df_blancos_otras_dilig.empty
            and proveedor in df_blancos_otras_dilig.columns):
        df_otras_prov = df_blancos_otras_dilig[["Proceso", proveedor]].rename(
            columns={proveedor: "% Diligenciado"}
        )

    return df_func_prov, df_nf_prov, df_otras_prov


def escribir_hojas_diligenciado_por_proveedor(writer, df_blancos_func_dilig, df_blancos_nf_dilig,
                                               df_blancos_otras_dilig, nombres_proveedores,
                                               filas_enteras_otras=None):
    """
    Escribe una hoja por proveedor (nombrada con el nombre del proveedor) con los
    mismos tres bloques de "% Diligenciado" (Requerimientos funcionales,
    Requerimientos no funcionales, Otras), pero mostrando únicamente los valores
    de ese proveedor en vez de una columna por proveedor.
    """
    if not nombres_proveedores:
        return

    cols_func = ["Proceso", "Columna F", "Columna G"]
    cols_nf = ["Proceso", "Columna D", "Columna E"]
    cols_otras = ["Proceso", "% Diligenciado"]

    nombres_hojas_usados = set(ws.title for ws in writer.book.worksheets)

    for proveedor in nombres_proveedores:
        df_func_prov, df_nf_prov, df_otras_prov = construir_bloques_diligenciado_proveedor(
            df_blancos_func_dilig, df_blancos_nf_dilig, df_blancos_otras_dilig, proveedor
        )

        hay_func = df_func_prov is not None and not df_func_prov.empty
        hay_nf = df_nf_prov is not None and not df_nf_prov.empty
        hay_otras = df_otras_prov is not None and not df_otras_prov.empty
        if not hay_func and not hay_nf and not hay_otras:
            continue

        sheet_name = _sanitizar_nombre_hoja(proveedor, nombres_hojas_usados)
        ws = writer.book.create_sheet(sheet_name)
        fila_actual = 1

        fila_actual = _escribir_bloque_blancos(
            ws, df_func_prov, "Requerimientos funcionales", cols_func, fila_actual, agrupar=True
        )
        fila_actual = _escribir_bloque_blancos(
            ws, df_nf_prov, "Requerimientos no funcionales", cols_nf, fila_actual, agrupar=False
        )
        if hay_otras:
            fila_actual = _escribir_bloque_blancos(
                ws, df_otras_prov, "Otras", cols_otras, fila_actual, agrupar=False,
                filas_enteras=filas_enteras_otras
            )

        n_cols_total = max(len(cols_func), len(cols_nf), len(cols_otras))
        for col_idx in range(1, n_cols_total + 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            max_len = 0
            for row_idx in range(1, fila_actual):
                cell = ws.cell(row=row_idx, column=col_idx)
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def escribir_hoja_blancos(writer, df_blancos_func, nombres_proveedores_func,
                           df_blancos_nf, nombres_proveedores_nf,
                           df_blancos_otras=None, nombres_proveedores_otras=None,
                           filas_enteras_otras=None, nombre_hoja="% en blanco"):
    """Escribe una hoja con tres bloques: requerimientos funcionales
    (columnas F/G), requerimientos no funcionales (columnas D/E) y "Otras"
    (filas adicionales definidas en OTRAS_FILAS_BLANCO_CONFIG, más
    las filas de conteo de registros definidas en OTRAS_CONTEO_REGISTROS_CONFIG).

    El parámetro `nombre_hoja` permite reutilizar esta función tanto para
    "% en blanco" como para "% Diligenciado" (misma estructura, valores invertidos)."""
    hay_func = df_blancos_func is not None and not df_blancos_func.empty
    hay_nf = df_blancos_nf is not None and not df_blancos_nf.empty
    hay_otras = df_blancos_otras is not None and not df_blancos_otras.empty
    if not hay_func and not hay_nf and not hay_otras:
        return

    columnas_func = ["Proceso"]
    for prov in nombres_proveedores_func:
        columnas_func.append(f"Columna F {prov}")
        columnas_func.append(f"Columna G {prov}")

    columnas_nf = ["Proceso"]
    for prov in nombres_proveedores_nf:
        columnas_nf.append(f"Columna D {prov}")
        columnas_nf.append(f"Columna E {prov}")

    columnas_otras = ["Proceso"] + (nombres_proveedores_otras or [])

    ws = writer.book.create_sheet(nombre_hoja)
    fila_actual = 1

    fila_actual = _escribir_bloque_blancos(
        ws, df_blancos_func, "Requerimientos funcionales", columnas_func, fila_actual, agrupar=True
    )
    fila_actual = _escribir_bloque_blancos(
        ws, df_blancos_nf, "Requerimientos no funcionales", columnas_nf, fila_actual, agrupar=False
    )
    if hay_otras:
        fila_actual = _escribir_bloque_blancos(
            ws, df_blancos_otras, "Otras", columnas_otras, fila_actual, agrupar=False,
            filas_enteras=filas_enteras_otras
        )

    n_cols_total = max(len(columnas_func), len(columnas_nf), len(columnas_otras))
    for col_idx in range(1, n_cols_total + 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, fila_actual):
            cell = ws.cell(row=row_idx, column=col_idx)
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


# =========================
# UI
# =========================
st.set_page_config(layout="wide")
st.title("Evaluador de Propuestas")

if st.sidebar.button("Reiniciar análisis"):
    st.session_state.clear()
    st.rerun()

with st.sidebar:
    st.session_state.setdefault("ni_peso_col_f", 100)
    st.session_state.setdefault("ni_peso_col_g", 100)
    st.session_state.setdefault("ni_si_estandar_pct",        100)
    st.session_state.setdefault("ni_si_componente_pct",       75)
    st.session_state.setdefault("ni_des_pct",                 50)
    st.session_state.setdefault("ni_ter_pct",                 50)
    st.session_state.setdefault("ni_no_pct",                   0)
    st.session_state.setdefault("ni_k_completa",             100)
    st.session_state.setdefault("ni_k_casi_completa",         75)
    st.session_state.setdefault("ni_k_parcialmente_completa", 50)
    st.session_state.setdefault("ni_k_incompleta",            25)
    st.session_state.setdefault("ni_k_totalmente_incompleta",  0)
    st.session_state.setdefault("ni_peso_total_cum",  100)
    st.session_state.setdefault("ni_peso_total_cal",  100)
    st.session_state.setdefault("ni_peso_alcance",    100)
    st.session_state.setdefault("ni_peso_metodologia",100)

    st.header("Pesos cumplimiento funcional")
    st.caption("Todos los pesos se ingresan de 0 a 100 (se convierten internamente a escala 0.0–1.0)")

    peso_col_f_pct = st.number_input(
        "Peso máximo columna F (cubrimiento) — rango: 0 a 100",
        min_value=0, max_value=100, step=5,
        key="ni_peso_col_f"
    )
    peso_col_g_pct = st.number_input(
        "Peso máximo columna G (inclusión) — rango: 0 a 100",
        min_value=0, max_value=100, step=5,
        key="ni_peso_col_g"
    )

    peso_col_f = peso_col_f_pct / 100
    peso_col_g = peso_col_g_pct / 100

    st.markdown("**Pesos cubrimiento (proporción del peso máximo col F):**")
    st.caption(
        "Respuestas válidas en col F: **SÍ (Estándar ERP)**, **Si (Componente Adicional)**, "
        "**DES**, **TER**, **NO**. Rango: 0 a 100"
    )

    _si_estandar_pct = st.number_input(
        "SÍ (Estándar ERP) — rango: 0 a 100, pred: 100",
        min_value=0, max_value=100, step=5,
        key="ni_si_estandar_pct"
    )
    _si_componente_pct = st.number_input(
        "Si (Componente Adicional) — rango: 0 a 100, pred: 75",
        min_value=0, max_value=100, step=5,
        key="ni_si_componente_pct"
    )
    _des_pct = st.number_input(
        "DES — rango: 0 a 100, pred: 50",
        min_value=0, max_value=100, step=5,
        key="ni_des_pct"
    )
    _ter_pct = st.number_input(
        "TER — rango: 0 a 100, pred: 50",
        min_value=0, max_value=100, step=5,
        key="ni_ter_pct"
    )
    _no_pct = st.number_input(
        "NO — rango: 0 a 100, pred: 0",
        min_value=0, max_value=100, step=5,
        key="ni_no_pct"
    )

    pesos_f = {
        "SÍ (ESTÁNDAR ERP)":        (_si_estandar_pct   / 100) * peso_col_f,
        "SI (COMPONENTE ADICIONAL)": (_si_componente_pct / 100) * peso_col_f,
        "DES":                       (_des_pct           / 100) * peso_col_f,
        "TER":                       (_ter_pct           / 100) * peso_col_f,
        "NO":                        (_no_pct            / 100) * peso_col_f,
        "VACIO": 0.0,
    }

    st.divider()
    st.markdown("**Pesos calidad (columna K):**")
    st.caption("Rango: 0 a 100")

    _k_completa = st.number_input(
        "COMPLETA (columna K) — rango: 0 a 100, pred: 100",
        min_value=0, max_value=100, step=5,
        key="ni_k_completa"
    )
    _k_casi_completa = st.number_input(
        "CASI COMPLETA (columna K) — rango: 0 a 100, pred: 75",
        min_value=0, max_value=100, step=5,
        key="ni_k_casi_completa"
    )
    _k_parcialmente_completa = st.number_input(
        "PARCIALMENTE COMPLETA (columna K) — rango: 0 a 100, pred: 50",
        min_value=0, max_value=100, step=5,
        key="ni_k_parcialmente_completa"
    )
    _k_incompleta = st.number_input(
        "INCOMPLETA (columna K) — rango: 0 a 100, pred: 25",
        min_value=0, max_value=100, step=5,
        key="ni_k_incompleta"
    )
    _k_totalmente_incompleta = st.number_input(
        "TOTALMENTE INCOMPLETA (columna K) — rango: 0 a 100, pred: 0",
        min_value=0, max_value=100, step=5,
        key="ni_k_totalmente_incompleta"
    )

    pesos_k = {
        "COMPLETA":               _k_completa               / 100,
        "CASI COMPLETA":          _k_casi_completa          / 100,
        "PARCIALMENTE COMPLETA":  _k_parcialmente_completa  / 100,
        "INCOMPLETA":             _k_incompleta             / 100,
        "TOTALMENTE INCOMPLETA":  _k_totalmente_incompleta  / 100,
        "VACIO": 0.0,
    }

    st.markdown("**Pesos totales del puntaje combinado:**")
    st.caption("Rango: 0 a 100")

    _peso_total_cumplimiento_pct = st.number_input(
        "Peso total cumplimiento — rango: 0 a 100, pred: 100",
        min_value=0, max_value=100, step=5,
        key="ni_peso_total_cum"
    )
    _peso_total_calidad_pct = st.number_input(
        "Peso total calidad — rango: 0 a 100, pred: 100",
        min_value=0, max_value=100, step=5,
        key="ni_peso_total_cal"
    )

    peso_total_cumplimiento = _peso_total_cumplimiento_pct / 100
    peso_total_calidad      = _peso_total_calidad_pct      / 100

    st.divider()
    st.markdown("**Peso alcance:**")
    st.caption("Rango: 0 a 100 — se aplica en la fórmula: (% SI × peso_cum + % calidad × peso_cal) × peso_alcance")
    _peso_alcance_pct = st.number_input(
        "Peso alcance — rango: 0 a 100, pred: 100",
        min_value=0, max_value=100, step=5,
        key="ni_peso_alcance"
    )
    peso_alcance = _peso_alcance_pct / 100

    st.markdown("**Peso metodología:**")
    st.caption("Rango: 0 a 100")
    _peso_metodologia_pct = st.number_input(
        "Peso metodología — rango: 0 a 100, pred: 100",
        min_value=0, max_value=100, step=5,
        key="ni_peso_metodologia"
    )
    peso_metodologia = _peso_metodologia_pct / 100

archivos = st.file_uploader("Sube archivos Excel", type=["xlsx"], accept_multiple_files=True)

if "archivos_cargados" not in st.session_state:
    st.session_state["archivos_cargados"] = False


# =========================
# PROCESAMIENTO
# =========================
if archivos and not st.session_state["archivos_cargados"]:
    data, data_k = [], []
    data_nf, data_k_nf = [], []
    detalles_globales, detalles_globales_k = {}, {}
    detalles_globales_nf, detalles_globales_k_nf = {}, {}
    data_experiencia_raw = []
    data_experiencia_oferente = []
    data_experiencia_oferente_raw = []
    data_alcance_servicios = []
    data_alcance_servicios_raw = []
    data_metodologia = []
    data_metodologia_raw = []
    data_info_solucion = []
    data_evolucion = []
    data_red_partners = []
    data_centros = []
    data_comunidades = []
    data_otras_blancos = {cfg["label"]: {} for cfg in OTRAS_FILAS_BLANCO_CONFIG}
    metadata_archivos = []
    nombres_proveedores = []

    for archivo in archivos:
        proveedor = Path(archivo.name).stem
        nombres_proveedores.append(proveedor)
        archivo_bytes = archivo.getvalue()
        tamano_kb = math.ceil(len(archivo_bytes) / 1024)

        fecha_modificacion = obtener_fecha_modificacion(archivo_bytes)

        metadata_archivos.append({
            "Nombre del archivo": archivo.name,
            "Tamaño (KB)": tamano_kb,
            "Fecha y hora de última modificación": fecha_modificacion,
        })

        with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(archivo_bytes)
            path = tmp.name

        (resultados, detalles, resultados_k, detalles_k,
         resultados_nf, detalles_nf, resultados_k_nf, detalles_k_nf) = analizar_archivo(
            path, pesos_f, peso_col_f, peso_col_g, pesos_k
        )

        for hoja, v in resultados.items():
            data.append({"Hoja": hoja, "Proveedor": proveedor, "Cumplimiento_%": v})
        for hoja, df_ in detalles.items():
            df_ = df_.copy(); df_["Proveedor"] = proveedor
            detalles_globales.setdefault(hoja, []).append(df_)
        for hoja, v in resultados_k.items():
            data_k.append({"Hoja": hoja, "Proveedor": proveedor, "Calidad_%": v})
        for hoja, df_ in detalles_k.items():
            df_ = df_.copy(); df_["Proveedor"] = proveedor
            detalles_globales_k.setdefault(hoja, []).append(df_)

        for hoja, v in resultados_nf.items():
            data_nf.append({"Hoja": hoja, "Proveedor": proveedor, "Cumplimiento_%": v})
        for hoja, df_ in detalles_nf.items():
            df_ = df_.copy(); df_["Proveedor"] = proveedor
            detalles_globales_nf.setdefault(hoja, []).append(df_)
        for hoja, v in resultados_k_nf.items():
            data_k_nf.append({"Hoja": hoja, "Proveedor": proveedor, "Calidad_%": v})
        for hoja, df_ in detalles_k_nf.items():
            df_ = df_.copy(); df_["Proveedor"] = proveedor
            detalles_globales_k_nf.setdefault(hoja, []).append(df_)

        wb_exp = openpyxl.load_workbook(path, data_only=True)

        for cfg in OTRAS_FILAS_BLANCO_CONFIG:
            blancos, total = analizar_blancos_rango(
                wb_exp, cfg["hoja_prefijo"], cfg["fila_inicio"], cfg["fila_fin"], cfg["columna"]
            )
            data_otras_blancos[cfg["label"]][proveedor] = (blancos, total)

        df_exp_raw = analizar_hoja_experiencia_raw(wb_exp, proveedor)
        if df_exp_raw is not None and not df_exp_raw.empty:
            data_experiencia_raw.append(df_exp_raw)

        df_exp_oferente = analizar_hoja_experiencia_oferente(wb_exp, proveedor)
        if df_exp_oferente is not None:
            data_experiencia_oferente.append(df_exp_oferente)

        df_exp_oferente_raw = analizar_hoja_experiencia_oferente_raw(wb_exp, proveedor)
        if df_exp_oferente_raw is not None and not df_exp_oferente_raw.empty:
            data_experiencia_oferente_raw.append(df_exp_oferente_raw)

        df_alcance = analizar_hoja_alcance_servicios(wb_exp, proveedor)
        if df_alcance is not None and not df_alcance.empty:
            data_alcance_servicios.append(df_alcance)

        df_alcance_raw_completo = analizar_hoja_alcance_servicios_raw(wb_exp, proveedor)
        if df_alcance_raw_completo is not None and not df_alcance_raw_completo.empty:
            data_alcance_servicios_raw.append(df_alcance_raw_completo)

        df_metodologia = analizar_hoja_metodologia(wb_exp, proveedor)
        if df_metodologia is not None:
            data_metodologia.append(df_metodologia)

        df_metodologia_raw = analizar_hoja_metodologia_raw(wb_exp, proveedor)
        if df_metodologia_raw is not None and not df_metodologia_raw.empty:
            data_metodologia_raw.append(df_metodologia_raw)

        df_info_sol = analizar_hoja_info_solucion(wb_exp, proveedor)
        if df_info_sol is not None and not df_info_sol.empty:
            data_info_solucion.append(df_info_sol)

        df_evol = analizar_hoja_evolucion(wb_exp, proveedor)
        if df_evol is not None and not df_evol.empty:
            data_evolucion.append(df_evol)

        df_red = analizar_hoja_ecosistema(wb_exp, proveedor)
        if df_red is not None and not df_red.empty:
            data_red_partners.append(df_red)

        df_centros = analizar_hoja_centros_id(wb_exp, proveedor)
        if df_centros is not None and not df_centros.empty:
            data_centros.append(df_centros)

        df_comunidades = analizar_hoja_comunidades(wb_exp, proveedor)
        if df_comunidades is not None and not df_comunidades.empty:
            data_comunidades.append(df_comunidades)

    df_final, df_total = construir_tablas_cumplimiento(data)
    df_final_k, df_total_k = construir_tablas_calidad(data_k) if data_k else (None, None)

    df_final_nf, df_total_nf = construir_tablas_cumplimiento(data_nf) if data_nf else (pd.DataFrame(), pd.DataFrame())
    df_final_k_nf, df_total_k_nf = construir_tablas_calidad(data_k_nf) if data_k_nf else (None, None)

    st.session_state.update({
        "df_final": df_final, "df_total": df_total,
        "df_final_k": df_final_k, "df_total_k": df_total_k,
        "df_final_nf": df_final_nf, "df_total_nf": df_total_nf,
        "df_final_k_nf": df_final_k_nf, "df_total_k_nf": df_total_k_nf,
        "detalles_globales": detalles_globales,
        "detalles_globales_k": detalles_globales_k,
        "detalles_globales_nf": detalles_globales_nf,
        "detalles_globales_k_nf": detalles_globales_k_nf,
        "data_experiencia_raw": data_experiencia_raw,
        "data_experiencia_oferente": data_experiencia_oferente,
        "data_experiencia_oferente_raw": data_experiencia_oferente_raw,
        "data_alcance_servicios": data_alcance_servicios,
        "data_alcance_servicios_raw": data_alcance_servicios_raw,
        "data_metodologia": data_metodologia,
        "data_metodologia_raw": data_metodologia_raw,
        "data_info_solucion": data_info_solucion,
        "data_evolucion": data_evolucion,
        "data_red_partners": data_red_partners,
        "data_centros": data_centros,
        "data_comunidades": data_comunidades,
        "data_otras_blancos": data_otras_blancos,
        "metadata_archivos": metadata_archivos,
        "nombres_proveedores": nombres_proveedores,
        "param_peso_col_f_raw": peso_col_f_pct,
        "param_peso_col_g_raw": peso_col_g_pct,
        "param_pesos_f_raw": {
            "SÍ (Estándar ERP)":        _si_estandar_pct,
            "Si (Componente Adicional)": _si_componente_pct,
            "DES":                       _des_pct,
            "TER":                       _ter_pct,
            "NO":                        _no_pct,
        },
        "param_pesos_k_raw": {
            "COMPLETA":              _k_completa,
            "CASI COMPLETA":         _k_casi_completa,
            "PARCIALMENTE COMPLETA": _k_parcialmente_completa,
            "INCOMPLETA":            _k_incompleta,
            "TOTALMENTE INCOMPLETA": _k_totalmente_incompleta,
        },
        "param_peso_total_cumplimiento_raw": _peso_total_cumplimiento_pct,
        "param_peso_total_calidad_raw":      _peso_total_calidad_pct,
        "archivos_cargados": True,
        "param_peso_alcance_raw": _peso_alcance_pct,
        "param_peso_metodologia_raw": _peso_metodologia_pct,
        "mostrar_total_func": False,
        "mostrar_total_nf": False,
        "mostrar_puntaje_func": False,
        "mostrar_puntaje_nf": False,
        "df_total_func_ponderado": None,
        "df_total_nf_ponderado": None,
        "df_puntaje_func": None,
        "df_total_puntaje_func": None,
        "df_puntaje_nf": None,
        "df_total_puntaje_nf": None,
    })


# =========================
# MOSTRAR
# =========================
if st.session_state["archivos_cargados"]:

    # ---- DETALLE DE HOJA ----
    if st.session_state.get("pagina_actual") == "detalle":
        hoja_d   = st.session_state.get("detalle_hoja")
        det_df   = st.session_state.get("detalle_df")
        det_df_k = st.session_state.get("detalle_df_k")

        st.subheader(f"Detalle de la hoja: {hoja_d}")

        def _df_to_excel(dfs):
            buf = BytesIO()
            with pd.ExcelWriter(buf, engine="openpyxl") as w:
                for sn, d in dfs.items():
                    d.to_excel(w, index=False, sheet_name=sn)
            return buf.getvalue()

        if det_df is not None:
            st.markdown("#### Cumplimiento por requerimiento")

            _df = det_df.copy()
            _df["Cumplimiento_%"] = _df["Peso_Total"] * 100

            _proveedores_det = _df["Proveedor"].unique().tolist()
            _lista_dfs_det = [_df[_df["Proveedor"] == p] for p in _proveedores_det if not _df[_df["Proveedor"] == p].empty]
            _orden_reqs = _orden_requerimientos(_lista_dfs_det)

            _pivot = _pivot_ordenado(_df, "Cumplimiento_%", _orden_reqs)
            _pivot_fmt = _pivot.copy()
            for c in _pivot_fmt.columns:
                if c not in ("ID", "Requerimiento"):
                    _pivot_fmt[c] = _pivot_fmt[c].apply(lambda x: f"{x:.2f}%")
            st.dataframe(_pivot_fmt, use_container_width=True)
            st.download_button("⬇️ Descargar cumplimiento por requerimiento",
                               _df_to_excel({"Cumplimiento por requerimiento": _pivot}),
                               file_name=f"cumplimiento_requerimiento_{hoja_d}.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               key="dl_det_pivot_cum")

            _resumen = det_df.groupby("Proveedor")["Peso_Total"].mean().mul(100).round(2).reset_index()
            _resumen_fmt = _resumen.copy()
            _resumen_fmt["Cumplimiento_%"] = _resumen_fmt["Peso_Total"].apply(lambda x: f"{x:.2f}%")
            _resumen_fmt = _resumen_fmt.drop(columns=["Peso_Total"])
            st.markdown("**Resumen cumplimiento**")
            st.dataframe(_resumen_fmt)
            st.download_button("⬇️ Descargar resumen cumplimiento",
                               _df_to_excel({"Resumen cumplimiento": _resumen.rename(columns={"Peso_Total": "Cumplimiento_%"})}),
                               file_name=f"resumen_cumplimiento_{hoja_d}.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               key="dl_det_resumen_cum")

        if det_df_k is not None:
            st.markdown("#### Calidad por requerimiento")

            _df_k = det_df_k.copy()
            _df_k["Calidad_%"] = _df_k["Peso_K"] * 100

            _proveedores_det_k = _df_k["Proveedor"].unique().tolist()
            _lista_dfs_det_k = [_df_k[_df_k["Proveedor"] == p] for p in _proveedores_det_k if not _df_k[_df_k["Proveedor"] == p].empty]
            _orden_reqs_k = _orden_requerimientos(_lista_dfs_det_k)

            _pivot_k = _pivot_ordenado(_df_k, "Calidad_%", _orden_reqs_k)
            _pivot_k_fmt = _pivot_k.copy()
            for c in _pivot_k_fmt.columns:
                if c not in ("ID", "Requerimiento"):
                    _pivot_k_fmt[c] = _pivot_k_fmt[c].apply(lambda x: f"{x:.2f}%")
            st.dataframe(_pivot_k_fmt, use_container_width=True)
            st.download_button("⬇️ Descargar calidad por requerimiento",
                               _df_to_excel({"Calidad por requerimiento": _pivot_k}),
                               file_name=f"calidad_requerimiento_{hoja_d}.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               key="dl_det_pivot_cal")

            _resumen_k = det_df_k.groupby("Proveedor")["Peso_K"].mean().mul(100).round(2).reset_index()
            _resumen_k_fmt = _resumen_k.copy()
            _resumen_k_fmt["Calidad_%"] = _resumen_k_fmt["Peso_K"].apply(lambda x: f"{x:.2f}%")
            _resumen_k_fmt = _resumen_k_fmt.drop(columns=["Peso_K"])
            st.markdown("**Resumen calidad**")
            st.dataframe(_resumen_k_fmt)
            st.download_button("⬇️ Descargar resumen calidad",
                               _df_to_excel({"Resumen calidad": _resumen_k.rename(columns={"Peso_K": "Calidad_%"})}),
                               file_name=f"resumen_calidad_{hoja_d}.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               key="dl_det_resumen_cal")

        _sheets_completo = {}
        if det_df is not None:
            _sheets_completo["Detalle cumplimiento"] = _pivot
            _sheets_completo["Resumen cumplimiento"] = _resumen.rename(columns={"Peso_Total": "Cumplimiento_%"})
        if det_df_k is not None:
            _sheets_completo["Detalle calidad"] = _pivot_k
            _sheets_completo["Resumen calidad"] = _resumen_k.rename(columns={"Peso_K": "Calidad_%"})
        st.divider()
        st.download_button("⬇️ Descargar todo (Excel completo)",
                           _df_to_excel(_sheets_completo),
                           file_name=f"detalle_{hoja_d}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           key="dl_det_completo")

        if st.button("← Volver"):
            st.session_state["pagina_actual"] = "principal"
            st.rerun()

        st.stop()
    # ---- FIN DETALLE ----

    df_final            = st.session_state["df_final"]
    df_total            = st.session_state["df_total"]
    df_final_k          = st.session_state["df_final_k"]
    df_total_k          = st.session_state["df_total_k"]
    df_final_nf         = st.session_state["df_final_nf"]
    df_total_nf         = st.session_state["df_total_nf"]
    df_final_k_nf       = st.session_state["df_final_k_nf"]
    df_total_k_nf       = st.session_state["df_total_k_nf"]
    detalles_globales      = st.session_state["detalles_globales"]
    detalles_globales_k    = st.session_state["detalles_globales_k"]
    detalles_globales_nf   = st.session_state["detalles_globales_nf"]
    detalles_globales_k_nf = st.session_state["detalles_globales_k_nf"]
    data_experiencia_raw      = st.session_state.get("data_experiencia_raw", [])
    data_experiencia_oferente = st.session_state.get("data_experiencia_oferente", [])
    data_experiencia_oferente_raw = st.session_state.get("data_experiencia_oferente_raw", [])
    data_alcance_servicios    = st.session_state.get("data_alcance_servicios", [])
    data_alcance_servicios_raw = st.session_state.get("data_alcance_servicios_raw", [])
    data_metodologia_raw      = st.session_state.get("data_metodologia_raw", [])
    data_info_solucion        = st.session_state.get("data_info_solucion", [])
    data_evolucion            = st.session_state.get("data_evolucion", [])
    data_red_partners         = st.session_state.get("data_red_partners", [])
    data_centros              = st.session_state.get("data_centros", [])
    data_comunidades          = st.session_state.get("data_comunidades", [])
    data_otras_blancos        = st.session_state.get("data_otras_blancos", {})
    metadata_archivos      = st.session_state.get("metadata_archivos", [])
    nombres_proveedores    = st.session_state.get("nombres_proveedores", [])

    # ---- FUNCIONAL ----
    st.subheader("Cumplimiento funcional")

    st.markdown("#### Cumplimiento por hoja")
    df_final_agrupado = agrupar_df_por_categoria(df_final)
    event = st.dataframe(formatear_porcentaje_df(df_final_agrupado), on_select="rerun", key="df_func")
    boton_descarga("⬇️ Descargar", {"Cumplimiento por hoja": df_final_agrupado}, "f_cumplimiento_hoja.xlsx", "dl_f_cum_hoja")
    if event.selection.rows:
        hoja = df_final_agrupado.iloc[event.selection.rows[0]]["Hoja"]
        if not es_fila_grupo(hoja):
            st.session_state["detalle_hoja"] = hoja
            st.session_state["detalle_df"] = pd.concat(detalles_globales[hoja])
            st.session_state["detalle_df_k"] = pd.concat(detalles_globales_k[hoja]) if hoja in detalles_globales_k else None
            st.session_state["pagina_actual"] = "detalle"
            st.rerun()

    st.markdown("#### Calidad por hoja")
    if df_final_k is not None:
        df_final_k_agrupado = agrupar_df_por_categoria(df_final_k)
        event_k = st.dataframe(formatear_porcentaje_df(df_final_k_agrupado), on_select="rerun", key="df_cal_func")
        boton_descarga("⬇️ Descargar", {"Calidad por hoja": df_final_k_agrupado}, "f_calidad_hoja.xlsx", "dl_f_cal_hoja")
        if event_k.selection.rows:
            hoja_k = df_final_k_agrupado.iloc[event_k.selection.rows[0]]["Hoja"]
            if not es_fila_grupo(hoja_k):
                st.session_state["detalle_hoja"] = hoja_k
                st.session_state["detalle_df"] = pd.concat(detalles_globales[hoja_k]) if hoja_k in detalles_globales else None
                st.session_state["detalle_df_k"] = pd.concat(detalles_globales_k[hoja_k])
                st.session_state["pagina_actual"] = "detalle"
                st.rerun()

    st.markdown("#### Pesos por hoja funcional")
    st.caption("Rango: 0 a 100 — indica el peso porcentual de cada hoja en el total funcional")
    hojas_func_list = df_final["Hoja"].tolist()

    for hoja_w in hojas_func_list:
        if f"peso_hoja_func_{hoja_w}" not in st.session_state:
            st.session_state[f"peso_hoja_func_{hoja_w}"] = 100.0

    pesos_hojas_func = {}

    def _render_pesos_hoja_func(hojas):
        for hoja_w in hojas:
            col_nombre, col_input = st.columns([2, 3])
            with col_nombre:
                st.markdown(f"<div style='padding-top:8px'>{hoja_w}</div>", unsafe_allow_html=True)
            with col_input:
                pesos_hojas_func[hoja_w] = st.number_input(
                    label=hoja_w,
                    min_value=0.0, max_value=100.0,
                    value=st.session_state.get(f"peso_hoja_func_{hoja_w}", 100),
                    step=0.1,
                    format="%.2f",
                    key=f"peso_hoja_func_{hoja_w}",
                    label_visibility="collapsed"
                )

    _hojas_por_grupo_func = {}
    for hoja_w in hojas_func_list:
        _hojas_por_grupo_func.setdefault(obtener_grupo_hoja(hoja_w), []).append(hoja_w)

    for grupo_nombre, _ in GRUPOS_HOJAS_FUNC:
        hojas_grupo = _hojas_por_grupo_func.get(grupo_nombre, [])
        if not hojas_grupo:
            continue
        st.markdown(f"**{grupo_nombre}**")
        _render_pesos_hoja_func(hojas_grupo)

    hojas_sin_grupo_func = _hojas_por_grupo_func.get(None, [])
    if hojas_sin_grupo_func:
        st.markdown("**Otras hojas**")
        _render_pesos_hoja_func(hojas_sin_grupo_func)

    _, col_btn_func, _ = st.columns([2, 1, 2])
    with col_btn_func:
        if st.button("Generar total funcional integrado", key="btn_total_func", use_container_width=True):
            pesos_actuales_func = {h: st.session_state.get(f"peso_hoja_func_{h}", 100) for h in hojas_func_list}
            _ptcum = st.session_state.get("ni_peso_total_cum", 100) / 100
            _ptcal = st.session_state.get("ni_peso_total_cal", 100) / 100
            df_integrado_func, df_total_integrado_func = construir_tabla_integrada(
                df_final, df_final_k, pesos_actuales_func, _ptcum, _ptcal
            )
            df_integrado_func = agrupar_df_por_categoria(df_integrado_func)
            st.session_state["df_integrado_func"] = df_integrado_func
            st.session_state["df_total_integrado_func"] = df_total_integrado_func
            st.session_state["mostrar_total_func"] = True
            st.session_state["snapshot_pesos_hojas_func"] = dict(pesos_actuales_func)

    if st.session_state.get("mostrar_total_func", False):
        st.markdown("#### Total funcional integrado (cumplimiento + calidad)")
        df_mostrar_func = st.session_state["df_integrado_func"]
        df_mostrar_total_func = st.session_state["df_total_integrado_func"]
        st.dataframe(formatear_porcentaje_df(df_mostrar_func), key="df_integrado_func_tbl")
        st.dataframe(formatear_porcentaje_df(df_mostrar_total_func), key="df_total_integrado_func_tbl")
        boton_descarga(
            "⬇️ Descargar",
            {"Total integrado funcional": df_mostrar_func, "TOTAL": df_mostrar_total_func},
            "f_total_integrado.xlsx",
            "dl_f_total_integrado"
        )

    # ---- NO FUNCIONAL ----
    st.subheader("Cumplimiento no funcional")

    st.markdown("#### Cumplimiento por hoja")
    event_nf = st.dataframe(formatear_porcentaje_df(df_final_nf), on_select="rerun", key="df_nofunc")
    boton_descarga("⬇️ Descargar", {"Cumplimiento por hoja": df_final_nf}, "nf_cumplimiento_hoja.xlsx", "dl_nf_cum_hoja")
    if event_nf.selection.rows:
        hoja_nf = df_final_nf.iloc[event_nf.selection.rows[0]]["Hoja"]
        st.session_state["detalle_hoja"] = hoja_nf
        st.session_state["detalle_df"] = pd.concat(detalles_globales_nf[hoja_nf])
        st.session_state["detalle_df_k"] = pd.concat(detalles_globales_k_nf[hoja_nf]) if hoja_nf in detalles_globales_k_nf else None
        st.session_state["pagina_actual"] = "detalle"
        st.rerun()

    st.markdown("#### Calidad por hoja (no funcional)")
    if df_final_k_nf is not None:
        event_k_nf = st.dataframe(formatear_porcentaje_df(df_final_k_nf), on_select="rerun", key="df_cal_nofunc")
        boton_descarga("⬇️ Descargar", {"Calidad por hoja": df_final_k_nf}, "nf_calidad_hoja.xlsx", "dl_nf_cal_hoja")
        if event_k_nf.selection.rows:
            hoja_k_nf = df_final_k_nf.iloc[event_k_nf.selection.rows[0]]["Hoja"]
            st.session_state["detalle_hoja"] = hoja_k_nf
            st.session_state["detalle_df"] = pd.concat(detalles_globales_nf[hoja_k_nf]) if hoja_k_nf in detalles_globales_nf else None
            st.session_state["detalle_df_k"] = pd.concat(detalles_globales_k_nf[hoja_k_nf])
            st.session_state["pagina_actual"] = "detalle"
            st.rerun()

    st.markdown("#### Pesos por hoja no funcional")
    st.caption("Rango: 0 a 100 — indica el peso porcentual de cada hoja en el total no funcional")
    hojas_nofunc_list = df_final_nf["Hoja"].tolist()

    for hoja_w in hojas_nofunc_list:
        if f"peso_hoja_nf_{hoja_w}" not in st.session_state:
            st.session_state[f"peso_hoja_nf_{hoja_w}"] = 100.0

    pesos_hojas_nf = {}
    for hoja_w in hojas_nofunc_list:
        col_nombre, col_input = st.columns([2, 3])
        with col_nombre:
            st.markdown(f"<div style='padding-top:8px'>{hoja_w}</div>", unsafe_allow_html=True)
        with col_input:
            pesos_hojas_nf[hoja_w] = st.number_input(
                label=hoja_w,
                min_value=0.0, max_value=100.0,
                value=st.session_state.get(f"peso_hoja_nf_{hoja_w}", 100.0),
                step=0.1,
                format="%.2f",
                key=f"peso_hoja_nf_{hoja_w}",
                label_visibility="collapsed"
            )

    _, col_btn_nf, _ = st.columns([2, 1, 2])
    with col_btn_nf:
        if st.button("Generar total no funcional integrado", key="btn_total_nf", use_container_width=True):
            pesos_actuales_nf = {h: st.session_state.get(f"peso_hoja_nf_{h}", 100) for h in hojas_nofunc_list}
            _ptcum = st.session_state.get("ni_peso_total_cum", 100) / 100
            _ptcal = st.session_state.get("ni_peso_total_cal", 100) / 100
            df_integrado_nf, df_total_integrado_nf = construir_tabla_integrada(
                df_final_nf, df_final_k_nf, pesos_actuales_nf, _ptcum, _ptcal
            )
            st.session_state["df_integrado_nf"] = df_integrado_nf
            st.session_state["df_total_integrado_nf"] = df_total_integrado_nf
            st.session_state["mostrar_total_nf"] = True
            st.session_state["snapshot_pesos_hojas_nf"] = dict(pesos_actuales_nf)

    if st.session_state.get("mostrar_total_nf", False):
        st.markdown("#### Total no funcional integrado (cumplimiento + calidad)")
        df_mostrar_nf = st.session_state["df_integrado_nf"]
        df_mostrar_total_nf = st.session_state["df_total_integrado_nf"]
        st.dataframe(formatear_porcentaje_df(df_mostrar_nf), key="df_integrado_nf_tbl")
        st.dataframe(formatear_porcentaje_df(df_mostrar_total_nf), key="df_total_integrado_nf_tbl")
        boton_descarga(
            "⬇️ Descargar",
            {"Total integrado no funcional": df_mostrar_nf, "TOTAL": df_mostrar_total_nf},
            "nf_total_integrado.xlsx",
            "dl_nf_total_integrado"
        )

    # ---- SOLIDEZ DEL FABRICANTE ----
    st.subheader("Solidez del fabricante")
    st.info(
        "La información consolidada de experiencia, diferenciadores técnicos, red de partners, mecanismos de soporte, mantenimiento, centros de I+D, comunidades colaborativas y "
        "ruta de evolución de la solución se encuentran en el reporte final."
    )

    # ---- CALIDAD DEL PROPONENTE ----
    st.subheader("Calidad del proponente")

    # ---- ALCANCE DE SERVICIOS ----
    st.markdown("#### Alcance de servicios")
    st.caption(
        "Fórmula: (% de SI × Peso total cumplimiento + % de calidad col E × Peso total calidad) × Peso alcance"
    )

    _ptcum_alc = st.session_state.get("ni_peso_total_cum", 100) / 100
    _ptcal_alc = st.session_state.get("ni_peso_total_cal", 100) / 100
    _pa_alc    = st.session_state.get("ni_peso_alcance",   100) / 100

    _pesos_k_alc = {
        "COMPLETA":               st.session_state.get("ni_k_completa",               100) / 100,
        "CASI COMPLETA":          st.session_state.get("ni_k_casi_completa",           75)  / 100,
        "PARCIALMENTE COMPLETA":  st.session_state.get("ni_k_parcialmente_completa",   50)  / 100,
        "INCOMPLETA":             st.session_state.get("ni_k_incompleta",              25)  / 100,
        "TOTALMENTE INCOMPLETA":  st.session_state.get("ni_k_totalmente_incompleta",    0)  / 100,
        "VACIO": 0.0,
    }

    df_alcance_tabla_fmt, df_alcance_tabla_raw = calcular_tabla_alcance(
        data_alcance_servicios, nombres_proveedores, _pesos_k_alc,
        _ptcum_alc, _ptcal_alc, _pa_alc,
    )

    if df_alcance_tabla_fmt is not None:
        st.dataframe(df_alcance_tabla_fmt, use_container_width=True, key="df_alcance_servicios")
        boton_descarga(
            "⬇️ Descargar alcance de servicios",
            {"Alcance de servicios": df_alcance_tabla_raw},
            "alcance_servicios.xlsx",
            "dl_alcance_servicios"
        )
    else:
        st.info("No se encontraron datos de alcance de servicios (hoja '7.').")

    st.markdown("#### Metodología Implementación")
    st.caption(
        "Fórmula: (% de SI × Peso total cumplimiento + % de calidad col E × Peso total calidad) × Peso metodología"
    )
    data_metodologia = st.session_state.get("data_metodologia", [])

    _ptcum_met = st.session_state.get("ni_peso_total_cum", 100) / 100
    _ptcal_met = st.session_state.get("ni_peso_total_cal", 100) / 100
    _pm_met    = st.session_state.get("ni_peso_metodologia", 100) / 100

    _pesos_k_met = {
        "COMPLETA":               st.session_state.get("ni_k_completa",               100) / 100,
        "CASI COMPLETA":          st.session_state.get("ni_k_casi_completa",           75)  / 100,
        "PARCIALMENTE COMPLETA":  st.session_state.get("ni_k_parcialmente_completa",   50)  / 100,
        "INCOMPLETA":             st.session_state.get("ni_k_incompleta",              25)  / 100,
        "TOTALMENTE INCOMPLETA":  st.session_state.get("ni_k_totalmente_incompleta",    0)  / 100,
        "VACIO": 0.0,
    }

    df_met_tabla_fmt, df_met_tabla_raw = calcular_tabla_metodologia(
        data_metodologia, nombres_proveedores, _pesos_k_met,
        _ptcum_met, _ptcal_met, _pm_met,
    )

    if df_met_tabla_fmt is not None:
        st.dataframe(df_met_tabla_fmt, use_container_width=True, key="df_metodologia")
        boton_descarga(
            "⬇️ Descargar metodología implementación",
            {"Metodología Implementación": df_met_tabla_raw},
            "metodologia_implementacion.xlsx",
            "dl_metodologia"
        )
    else:
        st.info("No se encontraron datos de metodología (hoja '8.').")

    st.info(
        "La información consolidada de experiencia del oferente "
        "se encuentra en el reporte final."
    )

    # ---- EXPORTAR EXCEL COMPLETO ----
    st.divider()

    _listo_func = st.session_state.get("mostrar_total_func", False)
    _listo_nf   = st.session_state.get("mostrar_total_nf", False)

    _pendientes = []
    if not _listo_func:
        _pendientes.append("**Total funcional integrado** — presiona «Generar total funcional integrado»")
    if not _listo_nf:
        _pendientes.append("**Total no funcional integrado** — presiona «Generar total no funcional integrado»")

    if _pendientes:
        st.info(
            "⚠️ Para habilitar la descarga del reporte completo, primero genera las siguientes tablas:\n\n"
            + "\n".join(f"- {p}" for p in _pendientes)
        )
        st.stop()

    pesos_hojas_func_reporte = st.session_state.get(
        "snapshot_pesos_hojas_func",
        {h: st.session_state.get(f"peso_hoja_func_{h}", 100) for h in df_final["Hoja"].tolist()}
    )
    pesos_hojas_nf_reporte = st.session_state.get(
        "snapshot_pesos_hojas_nf",
        {h: st.session_state.get(f"peso_hoja_nf_{h}", 100) for h in df_final_nf["Hoja"].tolist()}
    )

    _now_bogota      = datetime.now(tz=ZoneInfo("America/Bogota"))
    fecha_generacion = _now_bogota.strftime("%Y-%m-%d %H:%M:%S")
    _fecha_nombre    = _now_bogota.strftime("%d-%m-%Y-%H-%M-%S")
    nombre_reporte   = f"reporte-uniban-{_fecha_nombre}.xlsx"

    bloques_info = construir_hoja_info_analisis(
        fecha_generacion=fecha_generacion,
        peso_col_f=st.session_state.get("param_peso_col_f_raw", 100),
        peso_col_g=st.session_state.get("param_peso_col_g_raw", 100),
        pesos_f=st.session_state.get("param_pesos_f_raw", {
            "SÍ (Estándar ERP)": 100, "Si (Componente Adicional)": 75,
            "DES": 50, "TER": 50, "NO": 0,
        }),
        pesos_k=st.session_state.get("param_pesos_k_raw", {
            "COMPLETA": 100, "CASI COMPLETA": 75,
            "PARCIALMENTE COMPLETA": 50, "INCOMPLETA": 25, "TOTALMENTE INCOMPLETA": 0,
        }),
        peso_total_cumplimiento=st.session_state.get("param_peso_total_cumplimiento_raw", 100),
        peso_total_calidad=st.session_state.get("param_peso_total_calidad_raw", 100),
        pesos_hojas_func=pesos_hojas_func_reporte,
        pesos_hojas_nf=pesos_hojas_nf_reporte,
        metadata_archivos=metadata_archivos,
        peso_alcance=st.session_state.get("param_peso_alcance_raw", 100),
        peso_metodologia=st.session_state.get("param_peso_metodologia_raw", 100),
    )

    df_f_total_export            = st.session_state.get("df_total_func_ponderado") or df_total
    df_nf_total_export           = st.session_state.get("df_total_nf_ponderado") or df_total_nf
    df_puntaje_func_export       = st.session_state.get("df_puntaje_func")
    df_total_puntaje_func_export = st.session_state.get("df_total_puntaje_func")
    df_puntaje_nf_export         = st.session_state.get("df_puntaje_nf")
    df_total_puntaje_nf_export   = st.session_state.get("df_total_puntaje_nf")

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:

        def _write_pivot_block(ws, df, titulo, start_row):
            ws.cell(row=start_row, column=1, value=titulo).font = openpyxl.styles.Font(bold=True, size=11)
            start_row += 1
            for ci, col_name in enumerate(df.columns, start=1):
                ws.cell(row=start_row, column=ci, value=col_name).font = openpyxl.styles.Font(bold=True)
            start_row += 1
            for _, row in df.iterrows():
                for ci, val in enumerate(row, start=1):
                    ws.cell(row=start_row, column=ci, value=val)
                start_row += 1
            return start_row + 1

        for hoja_det, lista_dfs in detalles_globales.items():
            df_det = pd.concat(lista_dfs)
            df_det["Cumplimiento_%"] = df_det["Peso_Total"] * 100
            orden_reqs = _orden_requerimientos(lista_dfs)
            df_pivot_cum = _pivot_ordenado(df_det, "Cumplimiento_%", orden_reqs)

            df_pivot_cal = None
            if hoja_det in detalles_globales_k:
                lista_dfs_k = detalles_globales_k[hoja_det]
                df_det_k = pd.concat(lista_dfs_k)
                df_det_k["Calidad_%"] = df_det_k["Peso_K"] * 100
                orden_reqs_k = _orden_requerimientos(lista_dfs_k)
                df_pivot_cal = _pivot_ordenado(df_det_k, "Calidad_%", orden_reqs_k)

            sheet_name = f"Detalle de la hoja {hoja_det}"[:31]
            ws_det = writer.book.create_sheet(sheet_name)
            fila_ws = 1
            fila_ws = _write_pivot_block(ws_det, df_pivot_cum, "Cumplimiento por requerimiento", fila_ws)
            if df_pivot_cal is not None:
                fila_ws = _write_pivot_block(ws_det, df_pivot_cal, "Calidad por requerimiento", fila_ws)

        for hoja_det, lista_dfs in detalles_globales_nf.items():
            df_det = pd.concat(lista_dfs)
            df_det["Cumplimiento_%"] = df_det["Peso_Total"] * 100
            orden_reqs = _orden_requerimientos(lista_dfs)
            df_pivot_cum = _pivot_ordenado(df_det, "Cumplimiento_%", orden_reqs)

            df_pivot_cal = None
            if hoja_det in detalles_globales_k_nf:
                lista_dfs_k = detalles_globales_k_nf[hoja_det]
                df_det_k = pd.concat(lista_dfs_k)
                df_det_k["Calidad_%"] = df_det_k["Peso_K"] * 100
                orden_reqs_k = _orden_requerimientos(lista_dfs_k)
                df_pivot_cal = _pivot_ordenado(df_det_k, "Calidad_%", orden_reqs_k)

            sheet_name = f"Detalle de la hoja {hoja_det}"[:31]
            ws_det = writer.book.create_sheet(sheet_name)
            fila_ws = 1
            fila_ws = _write_pivot_block(ws_det, df_pivot_cum, "Cumplimiento por requerimiento", fila_ws)
            if df_pivot_cal is not None:
                fila_ws = _write_pivot_block(ws_det, df_pivot_cal, "Calidad por requerimiento", fila_ws)

        _safe_to_excel(agrupar_df_por_categoria(df_final),   writer, "F - Comparativo")
        _safe_to_excel(agrupar_df_por_categoria(df_final_k), writer, "F - Calidad por hoja")

        _df_integ_func       = st.session_state.get("df_integrado_func")
        _df_total_integ_func = st.session_state.get("df_total_integrado_func")
        if _df_integ_func is not None and not _df_integ_func.empty:
            _export_integ_func = pd.concat([_df_integ_func, _df_total_integ_func], ignore_index=True)
            _safe_to_excel(_export_integ_func, writer, "F - Total integrado")

        _safe_to_excel(df_puntaje_func_export,       writer, "F - Puntaje funcional")
        _safe_to_excel(df_total_puntaje_func_export, writer, "F - Total puntaje")

        _safe_to_excel(df_final_nf,   writer, "NF - Comparativo")
        _safe_to_excel(df_final_k_nf, writer, "NF - Calidad por hoja")

        _df_integ_nf       = st.session_state.get("df_integrado_nf")
        _df_total_integ_nf = st.session_state.get("df_total_integrado_nf")
        if _df_integ_nf is not None and not _df_integ_nf.empty:
            _export_integ_nf = pd.concat([_df_integ_nf, _df_total_integ_nf], ignore_index=True)
            _safe_to_excel(_export_integ_nf, writer, "NF - Total integrado")

        _safe_to_excel(df_puntaje_nf_export,       writer, "NF - Puntaje")
        _safe_to_excel(df_total_puntaje_nf_export, writer, "NF - Total puntaje")

        if data_experiencia_raw:
            df_exp_raw_export = pd.concat(data_experiencia_raw, ignore_index=True)
            _safe_to_excel(df_exp_raw_export, writer, "Exp - Fabricante completa")

        if data_info_solucion:
            df_info_sol_export = pd.concat(data_info_solucion, ignore_index=True)
            df_info_sol_pivot = pivotar_requerimiento_proveedor(df_info_sol_export, nombres_proveedores)
            _safe_to_excel(df_info_sol_pivot, writer, "Diferenciadores técnicos")

        if data_evolucion:
            df_evol_export = pd.concat(data_evolucion, ignore_index=True)
            df_evol_pivot = pivotar_requerimiento_proveedor(df_evol_export, nombres_proveedores)
            _safe_to_excel(df_evol_pivot, writer, "Ruta evolucion")

        if data_red_partners:
            df_red_export = pd.concat(data_red_partners, ignore_index=True)
            df_red_pivot = pivotar_requerimiento_proveedor(df_red_export, nombres_proveedores)
            _safe_to_excel(df_red_pivot, writer, "Red de partners")

        if data_centros:
            df_centros_export = pd.concat(data_centros, ignore_index=True)
            df_centros_pivot = pivotar_requerimiento_proveedor(df_centros_export, nombres_proveedores)
            _safe_to_excel(df_centros_pivot, writer, "Centros de (I+D)")

        if data_comunidades:
            df_comunidades_export = pd.concat(data_comunidades, ignore_index=True)
            df_comunidades_pivot = pivotar_requerimiento_proveedor(df_comunidades_export, nombres_proveedores)
            _safe_to_excel(df_comunidades_pivot, writer, "Comunidades colaborativas")

        if data_alcance_servicios_raw:
            df_alc_raw_export = pd.concat(data_alcance_servicios_raw, ignore_index=True)
            orden_servicios = list(dict.fromkeys(df_alc_raw_export["Servicio"].tolist()))

            df_alc_pivot_sino = (
                df_alc_raw_export[["Proveedor", "Servicio", "Incluido (SI/NO)"]]
                .pivot_table(index="Servicio", columns="Proveedor", values="Incluido (SI/NO)", aggfunc="first")
                .reindex(orden_servicios).reset_index()
            )
            df_alc_pivot_sino.columns.name = None
            cols_sino = ["Servicio"] + [p for p in nombres_proveedores if p in df_alc_pivot_sino.columns]
            df_alc_pivot_sino = df_alc_pivot_sino[cols_sino]

            df_alc_pivot_cal = (
                df_alc_raw_export[["Proveedor", "Servicio", "Calidad"]]
                .pivot_table(index="Servicio", columns="Proveedor", values="Calidad", aggfunc="first")
                .reindex(orden_servicios).reset_index()
            )
            df_alc_pivot_cal.columns.name = None
            cols_cal = ["Servicio"] + [p for p in nombres_proveedores if p in df_alc_pivot_cal.columns]
            df_alc_pivot_cal = df_alc_pivot_cal[cols_cal]

            ws_alc = writer.book.create_sheet("Alcance de servicios - completo")
            fila_ws = 1
            fila_ws = _write_pivot_block(ws_alc, df_alc_pivot_sino, "Incluido (SI/NO) por proveedor", fila_ws)
            fila_ws = _write_pivot_block(ws_alc, df_alc_pivot_cal,  "Calidad por proveedor",          fila_ws)
            for col in ws_alc.columns:
                max_len = max((len(str(c.value)) for c in col if c.value is not None), default=0)
                ws_alc.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

        _ptcum_rep = st.session_state.get("param_peso_total_cumplimiento_raw", 100) / 100
        _ptcal_rep = st.session_state.get("param_peso_total_calidad_raw",      100) / 100
        _pa_rep    = st.session_state.get("param_peso_alcance_raw",            100) / 100
        _pesos_k_rep = {
            "COMPLETA":               st.session_state.get("ni_k_completa",               100) / 100,
            "CASI COMPLETA":          st.session_state.get("ni_k_casi_completa",           75)  / 100,
            "PARCIALMENTE COMPLETA":  st.session_state.get("ni_k_parcialmente_completa",   50)  / 100,
            "INCOMPLETA":             st.session_state.get("ni_k_incompleta",              25)  / 100,
            "TOTALMENTE INCOMPLETA":  st.session_state.get("ni_k_totalmente_incompleta",    0)  / 100,
            "VACIO": 0.0,
        }
        _, df_alc_export_raw = calcular_tabla_alcance(
            data_alcance_servicios, nombres_proveedores, _pesos_k_rep,
            _ptcum_rep, _ptcal_rep, _pa_rep,
        )
        if df_alc_export_raw is not None:
            _safe_to_excel(df_alc_export_raw, writer, "Alcance de servicios")

        if data_metodologia_raw:
            df_met_raw_export = pd.concat(data_metodologia_raw, ignore_index=True)
            orden_elementos = list(dict.fromkeys(df_met_raw_export["Elemento de la Metodología"].tolist()))

            df_met_pivot_sino = (
                df_met_raw_export[["Proveedor", "Elemento de la Metodología", "Incluido (SI/NO)"]]
                .pivot_table(index="Elemento de la Metodología", columns="Proveedor", values="Incluido (SI/NO)", aggfunc="first")
                .reindex(orden_elementos).reset_index()
            )
            df_met_pivot_sino.columns.name = None
            cols_met_sino = ["Elemento de la Metodología"] + [p for p in nombres_proveedores if p in df_met_pivot_sino.columns]
            df_met_pivot_sino = df_met_pivot_sino[cols_met_sino]

            df_met_pivot_cal = (
                df_met_raw_export[["Proveedor", "Elemento de la Metodología", "Calidad"]]
                .pivot_table(index="Elemento de la Metodología", columns="Proveedor", values="Calidad", aggfunc="first")
                .reindex(orden_elementos).reset_index()
            )
            df_met_pivot_cal.columns.name = None
            cols_met_cal = ["Elemento de la Metodología"] + [p for p in nombres_proveedores if p in df_met_pivot_cal.columns]
            df_met_pivot_cal = df_met_pivot_cal[cols_met_cal]

            ws_met = writer.book.create_sheet("Metodologia - completa")
            fila_ws = 1
            fila_ws = _write_pivot_block(ws_met, df_met_pivot_sino, "Incluido (SI/NO) por proveedor", fila_ws)
            fila_ws = _write_pivot_block(ws_met, df_met_pivot_cal,  "Calidad por proveedor",          fila_ws)
            for col in ws_met.columns:
                max_len = max((len(str(c.value)) for c in col if c.value is not None), default=0)
                ws_met.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

        _ptcum_rep_met = st.session_state.get("param_peso_total_cumplimiento_raw", 100) / 100
        _ptcal_rep_met = st.session_state.get("param_peso_total_calidad_raw",      100) / 100
        _pm_rep        = st.session_state.get("param_peso_metodologia_raw",        100) / 100
        _pesos_k_met_rep = {
            "COMPLETA":               st.session_state.get("ni_k_completa",               100) / 100,
            "CASI COMPLETA":          st.session_state.get("ni_k_casi_completa",           75)  / 100,
            "PARCIALMENTE COMPLETA":  st.session_state.get("ni_k_parcialmente_completa",   50)  / 100,
            "INCOMPLETA":             st.session_state.get("ni_k_incompleta",              25)  / 100,
            "TOTALMENTE INCOMPLETA":  st.session_state.get("ni_k_totalmente_incompleta",    0)  / 100,
            "VACIO": 0.0,
        }
        data_metodologia_export = st.session_state.get("data_metodologia", [])
        _, df_met_export_raw = calcular_tabla_metodologia(
            data_metodologia_export, nombres_proveedores, _pesos_k_met_rep,
            _ptcum_rep_met, _ptcal_rep_met, _pm_rep,
        )
        if df_met_export_raw is not None:
            _safe_to_excel(df_met_export_raw, writer, "Metodologia Implementacion")

        if data_experiencia_oferente:
            _df_of_all = pd.concat(data_experiencia_oferente, ignore_index=True)
            _provs_of = list(_df_of_all["Proveedor"].unique())
            _pivot_of = (
                _df_of_all[_df_of_all["Sector/Industria"] != ""]
                .groupby(["Sector/Industria", "Proveedor"]).size()
                .unstack(fill_value=0)
                .reindex(columns=_provs_of, fill_value=0)
                .reset_index()
            )
            if data_experiencia_oferente_raw:
                df_exp_of_raw_export = pd.concat(data_experiencia_oferente_raw, ignore_index=True)
                _safe_to_excel(df_exp_of_raw_export, writer, "Exp - Oferente completa")
            _safe_to_excel(_pivot_of, writer, "Exp - Oferente por sector")

        # ---- % EN BLANCO (funcional F/G + no funcional D/E + Otras) — penúltima hoja del reporte ----
        df_blancos_func = construir_tabla_blancos_funcional(detalles_globales, nombres_proveedores)
        df_blancos_nf = construir_tabla_blancos_no_funcional(detalles_globales_nf, nombres_proveedores)
        df_blancos_otras = construir_tabla_blancos_otras(data_otras_blancos, nombres_proveedores)

        df_conteo_registros = construir_tabla_conteo_registros(
            OTRAS_CONTEO_REGISTROS_CONFIG,
            {
                "data_experiencia_raw": data_experiencia_raw,
                "data_experiencia_oferente_raw": data_experiencia_oferente_raw,
            },
            nombres_proveedores,
        )
        filas_enteras_otras = {cfg["label"] for cfg in OTRAS_CONTEO_REGISTROS_CONFIG}
        if df_conteo_registros is not None and not df_conteo_registros.empty:
            if df_blancos_otras is not None and not df_blancos_otras.empty:
                df_blancos_otras = pd.concat([df_blancos_otras, df_conteo_registros], ignore_index=True)
            else:
                df_blancos_otras = df_conteo_registros

        df_blancos_col_c = construir_tabla_blancos_columna_c(
            OTRAS_BLANCOS_COLUMNA_C_CONFIG,
            {
                "data_alcance_servicios": data_alcance_servicios,
                "data_metodologia": st.session_state.get("data_metodologia", []),
            },
            nombres_proveedores,
        )
        if df_blancos_col_c is not None and not df_blancos_col_c.empty:
            if df_blancos_otras is not None and not df_blancos_otras.empty:
                df_blancos_otras = pd.concat([df_blancos_otras, df_blancos_col_c], ignore_index=True)
            else:
                df_blancos_otras = df_blancos_col_c

        # ---- % DILIGENCIADO (equivalente a "% en blanco" pero invertido: 100 - % en blanco) ----
        df_blancos_func_dilig = invertir_porcentaje_df(df_blancos_func)
        df_blancos_nf_dilig = invertir_porcentaje_df(df_blancos_nf)
        df_blancos_otras_dilig = invertir_porcentaje_df(df_blancos_otras, filas_sin_invertir=filas_enteras_otras)

        # Nota: las hojas "% en blanco" y "% Diligenciado" ya NO se escriben en el
        # reporte final; ahora forman parte del "Reporte de validaciones" (ver más abajo).

        escribir_hoja_info_analisis(writer, bloques_info)

    # ---- REPORTE DE VALIDACIONES (hojas "% en blanco" + "% Diligenciado" + "Informacion de analisis") ----
    buffer_validaciones = BytesIO()
    with pd.ExcelWriter(buffer_validaciones, engine="openpyxl") as writer_val:
        escribir_hoja_blancos(
            writer_val, df_blancos_func, nombres_proveedores, df_blancos_nf, nombres_proveedores,
            df_blancos_otras, nombres_proveedores,
            filas_enteras_otras=filas_enteras_otras,
            nombre_hoja="% en blanco"
        )
        escribir_hojas_diligenciado_por_proveedor(
            writer_val, df_blancos_func_dilig, df_blancos_nf_dilig, df_blancos_otras_dilig,
            nombres_proveedores, filas_enteras_otras=filas_enteras_otras
        )
        escribir_hoja_info_analisis(writer_val, bloques_info)

    col_dl_reporte, col_dl_validaciones = st.columns(2)
    with col_dl_reporte:
        st.download_button(
            "⬇️ Descargar reporte completo Excel",
            buffer.getvalue(),
            file_name=nombre_reporte,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_reporte_completo"
        )
    with col_dl_validaciones:
        _nombre_reporte_val = f"reporte-validaciones-uniban-{_fecha_nombre}.xlsx"
        st.download_button(
            "⬇️ Descargar reporte de validaciones",
            buffer_validaciones.getvalue(),
            file_name=_nombre_reporte_val,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_reporte_validaciones"
        )